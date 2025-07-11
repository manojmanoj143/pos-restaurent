# -*- mode: python ; coding: utf-8 -*-
from flask import Flask, request, jsonify, send_from_directory, Response
from flask_cors import CORS
from pymongo import MongoClient
from pymongo.errors import ConnectionFailure
from bson.objectid import ObjectId
from datetime import datetime, timedelta, UTC
import os
import sys
import logging
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from werkzeug.utils import secure_filename
import openpyxl
from io import BytesIO
import schedule
import time
import threading
import waitress
import tenacity
import json
import secrets
import hashlib
from dotenv import load_dotenv
import bcrypt
import tempfile
import traceback
import uuid
from twilio.rest import Client
from twilio.base.exceptions import TwilioRestException
from dotenv import load_dotenv
from zoneinfo import ZoneInfo


# Load environment variables from .env file
load_dotenv()

# Set up logging configuration
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s %(levelname)s %(message)s',
    handlers=[logging.StreamHandler()]
)
logger = logging.getLogger(__name__)
logging.getLogger('pymongo').setLevel(logging.WARNING)
logging.getLogger('waitress').setLevel(logging.WARNING)

# Determine base directory
if getattr(sys, 'frozen', False):
    base_dir = os.path.dirname(sys.executable)
    if not os.access(base_dir, os.W_OK):
        logger.warning(f"No write permission in {base_dir}. Using temporary directory.")
        base_dir = tempfile.gettempdir()
else:
    base_dir = os.path.dirname(os.path.abspath(__file__))

# Use UPLOAD_FOLDER from environment variable if set (for production), else default to local path
UPLOAD_FOLDER = os.getenv('UPLOAD_FOLDER', os.path.join(base_dir, 'static', 'uploads'))

# Ensure necessary directories exist with permission handling
def create_directory(directory):
    try:
        if not os.path.exists(directory):
            os.makedirs(directory)
            logger.info(f"Created directory: {directory}")
    except PermissionError as e:
        logger.error(f"Permission denied creating directory {directory}: {str(e)}")
        fallback_dir = os.path.join(tempfile.gettempdir(), os.path.basename(directory))
        if not os.path.exists(fallback_dir):
            os.makedirs(fallback_dir)
            logger.info(f"Created fallback directory: {fallback_dir}")
        return fallback_dir
    except Exception as e:
        logger.error(f"Error creating directory {directory}: {str(e)}")
        raise
    return directory

# Initialize Flask app
app = Flask(__name__, static_folder=None, static_url_path=None)
STATIC_DIR = os.path.join(base_dir, 'dist')

# Create directories and set configurations
app.config['UPLOAD_FOLDER'] = create_directory(UPLOAD_FOLDER)
STATIC_DIR = create_directory(STATIC_DIR)

ALLOWED_IMAGE_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
ALLOWED_JSON_EXTENSIONS = {'json'}
MAX_BACKUPS = 5

# CORS configuration
CORS(app, resources={
    r"/api/*": {
        "origins": "*",
        "methods": ["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS"],
        "allow_headers": ["Content-Type", "Accept"]
    }
})

# MongoDB connection with retry logic
@tenacity.retry(
    wait=tenacity.wait_fixed(2),
    stop=tenacity.stop_after_attempt(5),
    reraise=True
)
def connect_to_mongodb():
    try:
        client = MongoClient("mongodb://localhost:27017/", serverSelectionTimeoutMS=5000)
        client.server_info()
        logger.info("Successfully connected to MongoDB")
        return client
    except ConnectionFailure as e:
        logger.error(f"Failed to connect to MongoDB: {str(e)}")
        raise

try:
    client = connect_to_mongodb()
    db = client['restaurant']
    items_collection = db['items']
    customers_collection = db['customers']
    sales_collection = db['sales']
    tables_collection = db['tables']
    users_collection = db['users']
    picked_up_collection = db['picked_up_items']
    opening_collection = db['pos_opening_entries']
    pos_closing_collection = db['pos_closing_entries']
    email_tokens_collection = db['email_tokens']
    settings_collection = db['system_settings']
    kitchens_collection = db['kitchens']
    item_groups_collection = db['item_groups']  
    kitchen_saved_collection = db['kitchen_saved']
    variants_collection = db['variants']
    employees_collection = db['employees']
    activeorders_collection = db['activeorders']
    tripreports_collection = db['tripreports']  # New collection for trip reports
    order_counters_collection = db['order_counters']
    email_settings_collection = db['email_settings']
    table_orders_collection = db['table_orders']  # New collection for table orders
    purchase_items_collection = db['purchase_items']
    suppliers_collection = db['suppliers']
    purchase_orders_collection = db['purchase_orders']
    purchase_receipts_collection = db['purchase_receipts']
    purchase_invoices_collection = db['purchase_invoices']
    
    
    

except Exception as e:
    logger.critical(f"Could not establish MongoDB connection: {str(e)}")
    sys.exit(1)

# Twilio setup
account_sid = 'AC05480a3334e70b80d9c62f379f45a7a0'
auth_token = '5c0f80b5626a34474c58da9ce4b1357c'
twilio_phone = '+12185304627'
# twilio_messaging_sid = 'MGf5700bd791ccba00ef2084cc78d41573'  # Uncomment if needed




# Check if all Twilio credentials are present
if not all([account_sid, auth_token, twilio_phone]):
    print("Error: Missing Twilio credentials")
    raise ValueError("Twilio credentials are not set")

# Initialize Twilio client
try:
    twilio_client = Client(account_sid, auth_token)
    print("Twilio client initialized successfully")
except Exception as e:
    print(f"Failed to initialize Twilio client: {str(e)}")
    raise



# Test users
TEST_USERS = [
    {
        "email": "admin@gmail.com",
        "password": "123",
        "phone_number": "1234567890",
        "role": "admin",
        "firstName": "admin",
        "company": "POS 8",
        "pos_profile": "POS-001",
        "status": "Active",
        "created_at": datetime.now(UTC).isoformat(),
        "is_test": True
    },
    {
        "email": "bearer@gmail.com",
        "password": "123",
        "phone_number": "0987654321",
        "role": "bearer",
        "firstName": "bearer",
        "company": "POS 8",
        "pos_profile": "POS-001",
        "status": "Active",
        "created_at": datetime.now(UTC).isoformat(),
        "is_test": True
    }
]

# Utility functions
def allowed_file(filename, allowed_extensions):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions

def convert_objectid_to_str(item):
    if isinstance(item, dict) and '_id' in item:
        item['_id'] = str(item['_id'])
    for field in ['addons', 'combos', 'variants']:
        if field in item and isinstance(item[field], list):
            for sub_item in item[field]:
                if '_id' in sub_item:
                    sub_item['_id'] = str(sub_item['_id'])
    return item

def handle_image_upload(file):
    if not file or not allowed_file(file.filename, ALLOWED_IMAGE_EXTENSIONS):
        logger.error(f"Invalid file or type: {file.filename if file else 'No file'}")
        return None
    filename = secure_filename(file.filename)
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if not os.path.exists(file_path):
        file.save(file_path)
        logger.info(f"New image uploaded: {filename}")
    else:
        logger.info(f"Using existing image: {filename}")
    return filename

def manage_backup_limit():
    backup_dir = app.config['UPLOAD_FOLDER']
    backups = [f for f in os.listdir(backup_dir) if f.startswith('backup_restaurant_data_') and f.endswith('.xlsx')]
    backups.sort(key=lambda x: os.path.getmtime(os.path.join(backup_dir, x)))
    while len(backups) >= MAX_BACKUPS:
        oldest_backup = backups.pop(0)
        os.remove(os.path.join(backup_dir, oldest_backup))
        logger.info(f"Removed oldest backup: {oldest_backup}")

def sanitize_image_fields(data):
    """Sanitize image fields to ensure only filenames are stored."""
    if 'image' in data and data['image']:
        data['image'] = data['image'].split('/')[-1] if '/' in data['image'] else data['image']
    if 'images' in data:
        data['images'] = [img.split('/')[-1] if '/' in img else img for img in data['images']]
    for addon in data.get('addons', []):
        if 'addon_image' in addon and addon['addon_image']:
            addon['addon_image'] = addon['addon_image'].split('/')[-1] if '/' in addon['addon_image'] else addon['addon_image']
        if 'spicy' in addon and addon['spicy']:
            if 'spicy_image' in addon['spicy'] and addon['spicy']['spicy_image']:
                addon['spicy']['spicy_image'] = addon['spicy']['spicy_image'].split('/')[-1] if '/' in addon['spicy']['spicy_image'] else addon['spicy']['spicy_image']
            if 'non_spicy_image' in addon['spicy'] and addon['spicy']['non_spicy_image']:
                addon['spicy']['non_spicy_image'] = addon['spicy']['non_spicy_image'].split('/')[-1] if '/' in addon['spicy']['non_spicy_image'] else addon['spicy']['non_spicy_image']
    for combo in data.get('combos', []):
        if 'combo_image' in combo and combo['combo_image']:
            combo['combo_image'] = combo['combo_image'].split('/')[-1] if '/' in combo['combo_image'] else combo['combo_image']
        if 'spicy' in combo and combo['spicy']:
            if 'spicy_image' in combo['spicy'] and combo['spicy']['spicy_image']:
                combo['spicy']['spicy_image'] = combo['spicy']['spicy_image'].split('/')[-1] if '/' in combo['spicy']['spicy_image'] else combo['spicy']['spicy_image']
            if 'non_spicy_image' in combo['spicy'] and combo['spicy']['non_spicy_image']:
                combo['spicy']['non_spicy_image'] = combo['spicy']['non_spicy_image'].split('/')[-1] if '/' in combo['spicy']['non_spicy_image'] else combo['spicy']['non_spicy_image']
    for variant in data.get('custom_variants', []):
        for subheading in variant.get('subheadings', []):
            if 'image' in subheading and subheading['image']:
                subheading['image'] = subheading['image'].split('/')[-1] if '/' in subheading['image'] else subheading['image']
    if 'spicy' in data and data['spicy']:
        if 'spicy_image' in data['spicy'] and data['spicy']['spicy_image']:
            data['spicy']['spicy_image'] = data['spicy']['spicy_image'].split('/')[-1] if '/' in data['spicy']['spicy_image'] else data['spicy']['spicy_image']
        if 'non_spicy_image' in data['spicy'] and data['spicy']['non_spicy_image']:
            data['spicy']['non_spicy_image'] = data['spicy']['non_spicy_image'].split('/')[-1] if '/' in data['spicy']['non_spicy_image'] else data['spicy']['non_spicy_image']
    return data

# System settings management
def get_system_settings():
    settings = settings_collection.find_one({"_id": "system_settings"})
    if not settings:
        default_settings = {
            "_id": "system_settings",
            "disableUserPassLogin": False,
            "allowLoginUsingMobileNumber": False,
            "allowLoginUsingUserName": True,
            "loginWithEmailLink": False,
            "sessionExpiry": "06:00",
            "documentShareKeyExpiry": 30,
            "denyMultipleSessions": False,
            "allowConsecutiveLoginAttempts": 5,
            "allowLoginAfterFail": 60,
            "enableTwoFactorAuth": False,
            "logoutOnPasswordReset": False,
            "forceUserToResetPassword": 0,
            "resetPasswordLinkExpiryDuration": "24:00",
            "passwordResetLimit": 3,
            "enablePasswordPolicy": False,
            "minimumPasswordScore": 2
        }
        settings_collection.insert_one(default_settings)
        return default_settings
    return settings

def save_system_settings(settings):
    settings["_id"] = "system_settings"
    settings_collection.replace_one({"_id": "system_settings"}, settings, upsert=True)

# API Routes
@app.route('/api/test', methods=['GET'])
def test_route():
    logger.info("Test route accessed")
    return jsonify({"message": "API is working"}), 200

@app.route('/api/import-mongodb', methods=['POST', 'OPTIONS'])
def import_mongodb():
    logger.info(f"Request received: {request.method} {request.path}")
    
    # Handle CORS preflight request
    if request.method == 'OPTIONS':
        response = jsonify({"success": True})
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Accept'
        return response, 200

    try:
        # Check if a file is included in the request
        if 'file' not in request.files:
            logger.error("No file part in request")
            return jsonify({"error": "No file uploaded"}), 400
        
        file = request.files['file']
        if file.filename == '':
            logger.error("No selected file")
            return jsonify({"error": "No selected file"}), 400
        
        # Validate file extension
        if not allowed_file(file.filename, ALLOWED_JSON_EXTENSIONS):
            logger.error(f"Invalid file type: {file.filename}")
            return jsonify({"error": "Only JSON files are allowed"}), 400

        filename = secure_filename(file.filename)
        
        # Extract collection name from filename (handle multiple dots)
        # Example: restaurant.item_groups.json -> item_groups
        collection_name = filename.rsplit('.', 1)[0].split('.')[-1]
        
        # Define valid collections
        valid_collections = [
            'users', 'tables', 'items', 'customers', 'sales',
            'picked_up_items', 'pos_opening_entries', 'pos_closing_entries',
            'kitchens', 'item_groups'  # Added item_groups
        ]
        
        # Validate collection name
        if not collection_name or collection_name not in valid_collections:
            logger.error(f"Invalid or unsupported collection name: {collection_name}")
            return jsonify({"error": f"Unsupported collection name: {collection_name}"}), 400

        target_collection = db[collection_name]
        
        # Read and parse JSON file
        data = json.loads(file.read().decode('utf-8'))
        if not isinstance(data, list):
            logger.error("JSON data must be an array")
            return jsonify({"error": "JSON data must be an array"}), 400

        inserted_count = 0
        for record in data:
            # Handle MongoDB ObjectId
            if '_id' in record and isinstance(record['_id'], dict) and '$oid' in record['_id']:
                record['_id'] = ObjectId(record['_id']['$oid'])
            
            # Add import timestamp
            record['imported_at'] = datetime.now(UTC).isoformat()

            # Define unique key for upsert based on collection
            unique_key = (
                {'_id': record.get('_id')} if '_id' in record else
                {'email': record.get('email')} if collection_name == 'users' else
                {'table_number': record.get('table_number')} if collection_name == 'tables' else
                {'item_name': record.get('item_name')} if collection_name == 'items' else
                {'phone_number': record.get('phone_number')} if collection_name == 'customers' else
                {'invoice_no': record.get('invoice_no')} if collection_name == 'sales' else
                {'customerName': record.get('customerName')} if collection_name == 'picked_up_items' else
                {'name': record.get('name')} if collection_name in ['pos_opening_entries', 'pos_closing_entries'] else
                {'kitchen_name': record.get('kitchen_name')} if collection_name == 'kitchens' else
                {'group_name': record.get('group_name')} if collection_name == 'item_groups' else
                {}
            )

            if not unique_key:
                logger.error(f"No unique key defined for record in collection {collection_name}")
                return jsonify({"error": f"No unique key defined for record in collection {collection_name}"}), 400

            # Perform upsert
            target_collection.replace_one(unique_key, record, upsert=True)
            inserted_count += 1

        logger.info(f"Imported {inserted_count} records into {collection_name}")
        return jsonify({"message": f"Successfully imported {inserted_count} records into {collection_name}"}), 200

    except json.JSONDecodeError as e:
        logger.error(f"Invalid JSON format in file {filename}: {str(e)}")
        return jsonify({"error": f"Invalid JSON format: {str(e)}"}), 400
    except Exception as e:
        logger.error(f"Error importing data: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/static/uploads/<filename>', methods=['GET'])
def serve_uploaded_image(filename):
    logger.debug(f"Serving image: {filename}")
    try:
        return send_from_directory(app.config['UPLOAD_FOLDER'], filename)
    except Exception as e:
        logger.error(f"Error serving image {filename}: {str(e)}")
        return jsonify({"error": "Image not found"}), 404

@app.route('/api/upload-image', methods=['POST', 'OPTIONS'])
def upload_image():
    if request.method == 'OPTIONS':
        response = jsonify({"success": True})
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        return response, 200
    try:
        if 'files' not in request.files:
            logger.error("No files part in request")
            return jsonify({"error": "No files provided"}), 400
        files = request.files.getlist('files')
        if not files or all(file.filename == '' for file in files):
            logger.error("No valid files selected")
            return jsonify({"error": "No valid files selected"}), 400
        urls = []
        for file in files:
            if file and allowed_file(file.filename, ALLOWED_IMAGE_EXTENSIONS):
                filename = handle_image_upload(file)
                if filename:
                    urls.append(filename)
                    logger.info(f"Uploaded image: {filename}")
                else:
                    logger.warning(f"Failed to upload image: {file.filename}")
            else:
                logger.warning(f"Invalid file type: {file.filename}")
        if not urls:
            return jsonify({"error": "No valid images uploaded"}), 400
        return jsonify({"urls": urls}), 200
    except Exception as e:
        logger.error(f"Error uploading images: {str(e)}")
        return jsonify({"error": f"Server error: {str(e)}"}), 500

@app.route('/api/delete-image/<filename>', methods=['DELETE', 'OPTIONS'])
def delete_image(filename):
    if request.method == 'OPTIONS':
        response = jsonify({"success": True})
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'DELETE, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        return response, 200
    try:
        item_id = request.args.get('item_id')
        field = request.args.get('field', 'image')
        valid_fields = {'image', 'images', 'addon_image', 'combo_image', 'variant_image'}
        if field not in valid_fields:
            logger.error(f"Invalid field specified: {field}")
            return jsonify({"error": f"Invalid field: {field}. Must be one of {valid_fields}"}), 400
        if not item_id:
            logger.error("Item ID is required for image deletion")
            return jsonify({"error": "Item ID is required"}), 400
        try:
            object_id = ObjectId(item_id)
        except Exception:
            logger.error(f"Invalid item ID: {item_id}")
            return jsonify({"error": "Invalid item ID"}), 400
        filename = secure_filename(filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        if field == 'images':
            result = items_collection.update_one(
                {"_id": object_id, "images": filename},
                {"$pull": {"images": filename}}
            )
            if result.modified_count == 0:
                logger.warning(f"Image {filename} not found in images array for item {item_id}")
                return jsonify({"error": "Image not found in item"}), 404
        elif field == 'image':
            result = items_collection.update_one(
                {"_id": object_id, "image": filename},
                {"$set": {"image": None}}
            )
            if result.modified_count == 0:
                logger.warning(f"Image {filename} not found in image field for item {item_id}")
                return jsonify({"error": "Image not found in item"}), 404
        elif field == 'addon_image':
            result = items_collection.update_one(
                {"_id": object_id, "addons.addon_image": filename},
                {"$set": {"addons.$[elem].addon_image": None}},
                array_filters=[{"elem.addon_image": filename}]
            )
            if result.modified_count == 0:
                logger.warning(f"Image {filename} not found in addons for item {item_id}")
                return jsonify({"error": "Image not found in addons"}), 404
        elif field == 'combo_image':
            result = items_collection.update_one(
                {"_id": object_id, "combos.combo_image": filename},
                {"$set": {"combos.$[elem].combo_image": None}},
                array_filters=[{"elem.combo_image": filename}]
            )
            if result.modified_count == 0:
                logger.warning(f"Image {filename} not found in combos for item {item_id}")
                return jsonify({"error": "Image not found in combos"}), 404
        elif field == 'variant_image':
            result = items_collection.update_one(
                {"_id": object_id, "variants.variant_image": filename},
                {"$set": {"variants.$[elem].variant_image": None}},
                array_filters=[{"elem.variant_image": filename}]
            )
            if result.modified_count == 0:
                logger.warning(f"Image {filename} not found in variants for item {item_id}")
                return jsonify({"error": "Image not found in variants"}), 404
        if os.path.exists(file_path):
            try:
                os.remove(file_path)
                logger.info(f"Image deleted from filesystem: {filename}")
            except PermissionError as e:
                logger.error(f"Permission denied deleting image {filename}: {str(e)}")
                return jsonify({"error": "Permission denied deleting image"}), 403
            except Exception as e:
                logger.error(f"Error deleting image {filename}: {str(e)}")
                return jsonify({"error": "Error deleting image"}), 500
        else:
            logger.warning(f"Image file not found on filesystem: {filename}")
        logger.info(f"Image {filename} deleted from {field} for item {item_id}")
        return jsonify({"message": "Image deleted successfully"}), 200
    except Exception as e:
        logger.error(f"Error deleting image {filename} for item {item_id}: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/login', methods=['POST'])
def login():
    try:
        data = request.get_json()
        settings = get_system_settings()
        identifier = data.get('identifier')
        password = data.get('password')
        login_type = data.get('type', '')
        if not identifier or not password:
            return jsonify({"message": "Identifier and password are required"}), 400

        user = None
        mobileOnly = settings.get('allowLoginUsingMobileNumber', False) and not settings.get('allowLoginUsingUserName', False) and not settings.get('loginWithEmailLink', False)
        mobileOrUsername = settings.get('allowLoginUsingMobileNumber', False) and settings.get('allowLoginUsingUserName', False) and not settings.get('loginWithEmailLink', False)
        allThree = settings.get('allowLoginUsingMobileNumber', False) and settings.get('allowLoginUsingUserName', False) and settings.get('loginWithEmailLink', False)

        if mobileOnly:
            user = users_collection.find_one({"phone_number": identifier})
        elif mobileOrUsername or login_type == 'mobile_or_username':
            user = users_collection.find_one({"phone_number": identifier}) or users_collection.find_one({"firstName": identifier})
        elif allThree or login_type == 'all':
            user = users_collection.find_one({"phone_number": identifier}) or users_collection.find_one({"firstName": identifier}) or users_collection.find_one({"email": identifier})
        else:
            return jsonify({"message": "No valid login method enabled"}), 403

        requires_opening_entry = False
        if user and bcrypt.checkpw(password.encode('utf-8'), user['password'].encode('utf-8')):
            last_opening_time = user.get('last_opening_entry_time')
            if last_opening_time:
                try:
                    last_opening = datetime.fromisoformat(last_opening_time.replace('Z', '+00:00'))
                    time_diff = datetime.now(ZoneInfo("UTC")) - last_opening
                    if time_diff.total_seconds() / 3600 >= 8:
                        requires_opening_entry = True
                except ValueError:
                    logger.warning(f"Invalid last_opening_entry_time format for user {identifier}")
                    requires_opening_entry = True
            else:
                requires_opening_entry = True
            response = {
                "message": "Login successful",
                "user": {
                    "id": str(user['_id']),
                    "username": user['firstName'],
                    "role": user['role'],
                    "email": user.get('email', ''),
                    "phone_number": user.get('phone_number', ''),
                    "pos_profile": user.get('pos_profile', 'POS-001'),
                    "company": user.get('company', 'POS 8'),
                    "is_test": user.get('is_test', False)
                },
                "requires_opening_entry": requires_opening_entry
            }
            logger.info(f"User logged in: {identifier}, role: {user['role']}, requires_opening_entry: {requires_opening_entry}")
            return jsonify(response), 200

        # Check for test user
        test_user = next((u for u in TEST_USERS if ((u['firstName'] == identifier or u['phone_number'] == identifier or u['email'] == identifier) and u['password'] == password)), None)
        if test_user:
            # Check if test user already exists in the database
            existing_user = users_collection.find_one({"email": test_user['email']})
            if not existing_user:
                # Hash the test user's password and insert into the database
                hashed_password = bcrypt.hashpw(test_user['password'].encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
                new_user = {
                    "email": test_user['email'],
                    "password": hashed_password,
                    "phone_number": test_user['phone_number'],
                    "role": test_user['role'],
                    "firstName": test_user['firstName'],
                    "company": test_user['company'],
                    "pos_profile": test_user['pos_profile'],
                    "status": test_user['status'],
                    "created_at": test_user['created_at'],
                    "is_test": test_user['is_test']
                }
                result = users_collection.insert_one(new_user)
                logger.info(f"Test user {test_user['email']} added to database with ID: {str(result.inserted_id)}")
                user_id = str(result.inserted_id)
            else:
                user_id = str(existing_user['_id'])

            response = {
                "message": "Login successful",
                "user": {
                    "id": user_id,
                    "username": test_user['firstName'],
                    "role": test_user['role'],
                    "email": test_user.get('email', ''),
                    "phone_number": test_user.get('phone_number', ''),
                    "pos_profile": test_user.get('pos_profile', 'POS-001'),
                    "company": test_user.get('company', 'POS 8'),
                    "is_test": True
                },
                "requires_opening_entry": False
            }
            logger.info(f"Login with test credentials: {identifier}, role: {test_user['role']}")
            return jsonify(response), 200

        logger.warning(f"Invalid login attempt: {identifier}")
        return jsonify({"message": "Invalid credentials"}), 401
    except Exception as e:
        logger.error(f"Login failed: {str(e)}")
        return jsonify({"message": f"Login failed: {str(e)}"}), 500

@app.route('/api/request-email-login', methods=['POST'])
def request_email_login():
    try:
        settings = get_system_settings()
        if not settings.get('loginWithEmailLink', False):
            return jsonify({"message": "Email login is not enabled"}), 403
        data = request.get_json()
        email = data.get('email')
        if not email:
            return jsonify({"message": "Email is required"}), 400
        user = users_collection.find_one({"email": email})
        if not user:
            return jsonify({"message": "User not found"}), 404
        token = secrets.token_urlsafe(32)
        token_hash = hashlib.sha256(token.encode()).hexdigest()
        expiry = datetime.now(ZoneInfo("UTC")) + timedelta(hours=24)
        email_tokens_collection.insert_one({
            "email": email,
            "token_hash": token_hash,
            "expiry": expiry.isoformat(),
            "used": False,
            "created_at": datetime.now(ZoneInfo("UTC")).isoformat()
        })
        login_link = f"http://localhost:5000/api/verify-email-login?token={token}"
        html_content = f"""
        <h1>Login to Your Account</h1>
        <p>Click the link below to log in:</p>
        <a href="{login_link}">Login Now</a>
        <p>This link expires in 24 hours.</p>
        """
        msg = MIMEMultipart('alternative')
        msg['From'] = FROM_EMAIL
        msg['To'] = email
        msg['Subject'] = "Your Login Link"
        msg.attach(MIMEText(html_content, 'html'))
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(EMAIL_USER, EMAIL_PASS)
            server.send_message(msg)
        logger.info(f"Email login link sent to: {email}")
        return jsonify({"message": "Login link sent to your email"}), 200
    except Exception as e:
        logger.error(f"Error sending email login link: {str(e)}")
        return jsonify({"message": f"Failed: {str(e)}"}), 500

@app.route('/api/verify-email-login', methods=['GET'])
def verify_email_login():
    try:
        settings = get_system_settings()
        if not settings.get('loginWithEmailLink', False):
            return jsonify({"message": "Email login is not enabled"}), 403
        token = request.args.get('token')
        if not token:
            return jsonify({"message": "Token is required"}), 400
        token_hash = hashlib.sha256(token.encode()).hexdigest()
        token_doc = email_tokens_collection.find_one({"token_hash": token_hash})
        if not token_doc or token_doc['used'] or datetime.fromisoformat(token_doc['expiry']) < datetime.now(ZoneInfo("UTC")):
            return jsonify({"message": "Invalid or expired token"}), 401
        user = users_collection.find_one({"email": token_doc['email']})
        if not user:
            return jsonify({"message": "User not found"}), 404
        email_tokens_collection.update_one({"_id": token_doc['_id']}, {"$set": {"used": True}})
        response = {
            "message": "Login successful",
            "user": {
                "id": str(user['_id']),
                "username": user['firstName'],
                "role": user['role'],
                "email": user.get('email', ''),
                "phone_number": user.get('phone_number', ''),
                "pos_profile": user.get('pos_profile', 'POS-001'),
                "company": user.get('company', 'POS 8'),
                "is_test": user.get('is_test', False)
            }
        }
        logger.info(f"Email login successful for: {user['email']}")
        return jsonify(response), 200
    except Exception as e:
        logger.error(f"Error verifying email login: {str(e)}")
        return jsonify({"message": f"Failed: {str(e)}"}), 500

@app.route('/api/register', methods=['POST'])
def register():
    try:
        data = request.get_json()
        email = data.get('email')
        password = data.get('password')
        role = data.get('role')
        firstName = data.get('firstName')
        phone_number = data.get('phoneNumber')
        company = data.get('company', 'POS 8')
        if not email or not password or not role or not firstName or not phone_number:
            logger.error("Missing required fields in registration")
            return jsonify({"message": "Email, password, role, firstName, and phoneNumber are required"}), 400
        if email in [u['email'] for u in TEST_USERS]:
            logger.warning(f"Registration attempt with test email: {email}")
            return jsonify({"message": "Cannot register with test credentials"}), 400
        if users_collection.find_one({"email": email}):
            logger.warning(f"Registration attempt with existing email: {email}")
            return jsonify({"message": "Email already registered"}), 400
        if users_collection.find_one({"phone_number": phone_number}):
            logger.warning(f"Registration attempt with existing phone number: {phone_number}")
            return jsonify({"message": "Phone number already registered"}), 400
        hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        new_user = {
            "email": email,
            "password": hashed_password,
            "role": role,
            "firstName": firstName,
            "phone_number": phone_number,
            "company": company,
            "pos_profile": "POS-001",
            "status": "Active",
            "created_at": datetime.now(ZoneInfo("UTC")).isoformat()
        }
        result = users_collection.insert_one(new_user)
        logger.info(f"User registered: {email}, role: {role}")
        return jsonify({
            "message": "Registration successful",
            "user": {
                "id": str(result.inserted_id),
                "email": email,
                "role": role,
                "firstName": firstName,
                "phone_number": phone_number,
                "company": company
            }
        }), 201
    except Exception as e:
        logger.error(f"Registration failed: {str(e)}")
        return jsonify({"message": f"Registration failed: {str(e)}"}), 500

@app.route('/api/users', methods=['GET'])
def get_users():
    try:
        users = list(users_collection.find({}, {"password": 0}))
        users_list = [convert_objectid_to_str(user) for user in users]
        logger.info(f"Fetched {len(users_list)} users from MongoDB")
        return jsonify(users_list), 200
    except Exception as e:
        logger.error(f"Error fetching users: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/users/<email>', methods=['DELETE'])
def delete_user(email):
    try:
        if email in [u['email'] for u in TEST_USERS]:
            logger.warning(f"Attempt to delete test user: {email}")
            return jsonify({"message": "Cannot delete test users"}), 400
        result = users_collection.delete_one({"email": email})
        if result.deleted_count == 0:
            logger.warning(f"User not found for deletion: {email}")
            return jsonify({"message": "User not found"}), 404
        logger.info(f"User deleted: {email}")
        return jsonify({"message": "User deleted successfully"}), 200
    except Exception as e:
        logger.error(f"Error deleting user {email}: {str(e)}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/settings', methods=['GET'])
def get_settings():
    try:
        settings = get_system_settings()
        logger.info("Fetched system settings")
        return jsonify(settings), 200
    except Exception as e:
        logger.error(f"Error fetching settings: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/settings', methods=['POST'])
def update_settings():
    try:
        data = request.get_json()
        save_system_settings(data)
        logger.info("System settings updated")
        return jsonify({"message": "Settings updated successfully"}), 200
    except Exception as e:
        logger.error(f"Error updating settings: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/items', methods=['GET'])
def get_items():
    try:
        items = list(items_collection.find())
        items_list = []
        current_time = datetime.now(UTC)
        for item in items:
            item = convert_objectid_to_str(item)
            is_offer_active = False
            if 'offer_start_time' in item and item['offer_start_time'] and 'offer_end_time' in item and item['offer_end_time']:
                try:
                    offer_start_time = datetime.fromisoformat(str(item['offer_start_time']).replace('Z', '+00:00'))
                    offer_end_time = datetime.fromisoformat(str(item['offer_end_time']).replace('Z', '+00:00'))
                    if offer_start_time <= current_time <= offer_end_time:
                        is_offer_active = True
                        logger.debug(f"Offer active for item {item['_id']}: {offer_start_time} to {offer_end_time}")
                    else:
                        logger.debug(f"Offer inactive for item {item['_id']}: {offer_start_time} to {offer_end_time}")
                except (ValueError, TypeError) as e:
                    logger.warning(f"Invalid offer time format for item {item['_id']}: {str(e)}")
                    items_collection.update_one(
                        {'_id': ObjectId(item['_id'])},
                        {'$unset': {'offer_price': "", 'offer_start_time': "", 'offer_end_time': ""}}
                    )
                    item.pop('offer_price', None)
                    item.pop('offer_start_time', None)
                    item.pop('offer_end_time', None)
            elif 'offer_end_time' in item and item['offer_end_time']:
                try:
                    offer_end_time = datetime.fromisoformat(str(item['offer_end_time']).replace('Z', '+00:00'))
                    if current_time <= offer_end_time:
                        is_offer_active = True
                        logger.debug(f"Offer active for item {item['_id']}: ends at {offer_end_time}")
                    else:
                        logger.debug(f"Offer expired for item {item['_id']}: ended at {offer_end_time}")
                except (ValueError, TypeError) as e:
                    logger.warning(f"Invalid offer_end_time format for item {item['_id']}: {str(e)}")
                    items_collection.update_one(
                        {'_id': ObjectId(item['_id'])},
                        {'$unset': {'offer_price': "", 'offer_start_time': "", 'offer_end_time': ""}}
                    )
                    item.pop('offer_price', None)
                    item.pop('offer_start_time', None)
                    item.pop('offer_end_time', None)
            if not is_offer_active:
                item.pop('offer_price', None)
                item.pop('offer_start_time', None)
                item.pop('offer_end_time', None)
            if 'image' in item and item['image']:
                item['image'] = f"/static/uploads/{item['image']}"
            for addon in item.get("addons", []):
                if 'addon_image' in addon and addon['addon_image']:
                    addon['addon_image'] = f"/static/uploads/{addon['addon_image']}"
            for combo in item.get("combos", []):
                if 'combo_image' in combo and combo['combo_image']:
                    combo['combo_image'] = f"/static/uploads/{combo['combo_image']}"
            for variant in item.get("variants", []):
                if 'variant_image' in variant and variant['variant_image']:
                    variant['variant_image'] = f"/static/uploads/{variant['variant_image']}"
            items_list.append(item)
        logger.info(f"Fetched {len(items_list)} items")
        return jsonify(items_list), 200
    except Exception as e:
        logger.error(f"Error fetching items: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/items/<identifier>', methods=['GET'])
def get_item(identifier):
    try:
        item = None
        try:
            object_id = ObjectId(identifier)
            item = items_collection.find_one({'_id': object_id})
        except Exception:
            item = items_collection.find_one({'item_name': identifier})
        if not item:
            logger.warning(f"Item not found: {identifier}")
            return jsonify({"error": "Item not found"}), 404
        item = convert_objectid_to_str(item)
        current_time = datetime.now(UTC)
        is_offer_active = False
        if 'offer_start_time' in item and item['offer_start_time'] and 'offer_end_time' in item and item['offer_end_time']:
            try:
                offer_start_time = datetime.fromisoformat(str(item['offer_start_time']).replace('Z', '+00:00'))
                offer_end_time = datetime.fromisoformat(str(item['offer_end_time']).replace('Z', '+00:00'))
                if offer_start_time <= current_time <= offer_end_time:
                    is_offer_active = True
                    logger.debug(f"Offer active for item {item['_id']}: {offer_start_time} to {offer_end_time}")
                else:
                    logger.debug(f"Offer inactive for item {item['_id']}: {offer_start_time} to {offer_end_time}")
            except (ValueError, TypeError) as e:
                logger.warning(f"Invalid offer time format for item {item['_id']}: {str(e)}")
                items_collection.update_one(
                    {'_id': ObjectId(item['_id'])},
                    {'$unset': {'offer_price': "", 'offer_start_time': "", 'offer_end_time': ""}}
                )
                item.pop('offer_price', None)
                item.pop('offer_start_time', None)
                item.pop('offer_end_time', None)
        elif 'offer_end_time' in item and item['offer_end_time']:
            try:
                offer_end_time = datetime.fromisoformat(str(item['offer_end_time']).replace('Z', '+00:00'))
                if current_time <= offer_end_time:
                    is_offer_active = True
                    logger.debug(f"Offer active for item {item['_id']}: ends at {offer_end_time}")
                else:
                    logger.debug(f"Offer expired for item {item['_id']}: ended at {offer_end_time}")
            except (ValueError, TypeError) as e:
                logger.warning(f"Invalid offer_end_time for item {item['_id']}: {str(e)}")
                items_collection.update_one(
                    {'_id': ObjectId(item['_id'])},
                    {'$unset': {'offer_price': "", 'offer_start_time': "", 'offer_end_time': ""}}
                )
                item.pop('offer_price', None)
                item.pop('offer_start_time', None)
                item.pop('offer_end_time', None)
        if not is_offer_active:
            item.pop('offer_price', None)
            item.pop('offer_start_time', None)
            item.pop('offer_end_time', None)
        if 'image' in item and item['image']:
            item['image'] = f"/static/uploads/{item['image']}"
        for addon in item.get('addons', []):
            if 'addon_image' in addon and addon['addon_image']:
                addon['addon_image'] = f"/static/uploads/{addon['addon_image']}"
        for combo in item.get('combos', []):
            if 'combo_image' in combo and combo['combo_image']:
                combo['combo_image'] = f"/static/uploads/{combo['combo_image']}"
        logger.info(f"Fetched item: {identifier}")
        return jsonify(item), 200
    except Exception as e:
        logger.error(f"Error fetching item {identifier}: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/items', methods=['POST'])
def create_item():
    try:
        data = request.json
        if not data:
            logger.error("No data provided for item creation")
            return jsonify({"error": "No data provided"}), 400
        required_fields = ['item_name', 'item_code', 'item_group', 'price_list_rate']
        for field in required_fields:
            if field not in data or not data[field]:
                logger.error(f"Missing or empty required field: {field}")
                return jsonify({"error": f"Missing or empty required field: {field}"}), 400
        if 'offer_start_time' in data and data['offer_start_time'] and 'offer_end_time' in data and data['offer_end_time']:
            try:
                offer_start_time = datetime.fromisoformat(str(data['offer_start_time']).replace('Z', '+00:00'))
                offer_end_time = datetime.fromisoformat(str(data['offer_end_time']).replace('Z', '+00:00'))
                if offer_start_time >= offer_end_time:
                    logger.error("offer_start_time must be before offer_end_time")
                    return jsonify({"error": "Offer start time must be before offer end time"}), 400
            except (ValueError, TypeError) as e:
                logger.error(f"Invalid offer time format: {str(e)}")
                return jsonify({"error": f"Invalid offer time format: {str(e)}"}), 400
        data = sanitize_image_fields(data)
        data.setdefault('custom_addon_applicable', False)
        data.setdefault('custom_combo_applicable', False)
        data.setdefault('custom_total_calories', 0)
        data.setdefault('custom_total_protein', 0)
        data.setdefault('kitchen', "")
        data.setdefault('has_variant_pricing', False)
        data.setdefault('variant_prices', {"small_price": 0, "medium_price": 0, "large_price": 0})
        data.setdefault('variant_quantities', {"small_quantity": 0, "medium_quantity": 0, "large_quantity": 0})
        data.setdefault('sold_quantities', {"small_sold": 0, "medium_sold": 0, "large_sold": 0})
        data.setdefault('ice_preference', "without_ice")
        data.setdefault('ice_price', 0)
        data.setdefault('addons', [])
        data.setdefault('combos', [])
        data.setdefault('ingredients', [])
        data.setdefault('variants', [])
        data.setdefault('custom_variants', [])  # Added for custom variants
        data['created_at'] = datetime.now(UTC).isoformat()
        item_id = items_collection.insert_one(data).inserted_id
        logger.info(f"Item created with ID: {item_id}")
        return jsonify({'message': 'Item created successfully!', 'id': str(item_id)}), 201
    except Exception as e:
        logger.error(f"Error creating item: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/items/<item_id>', methods=['PUT'])
def update_item(item_id):
    try:
        item_data = request.get_json()
        if not item_data:
            logger.error("No data provided for item update")
            return jsonify({"error": "No data provided"}), 400
        try:
            object_id = ObjectId(item_id)
        except Exception:
            logger.error(f"Invalid item ID: {item_id}")
            return jsonify({"error": "Invalid item ID"}), 400
        if '_id' in item_data:
            del item_data['_id']
        if 'offer_start_time' in item_data and item_data['offer_start_time'] and 'offer_end_time' in item_data and item_data['offer_end_time']:
            try:
                offer_start_time = datetime.fromisoformat(str(item_data['offer_start_time']).replace('Z', '+00:00'))
                offer_end_time = datetime.fromisoformat(str(item_data['offer_end_time']).replace('Z', '+00:00'))
                if offer_start_time >= offer_end_time:
                    logger.error("offer_start_time must be before offer_end_time")
                    return jsonify({"error": "Offer start time must be before offer end time"}), 400
            except (ValueError, TypeError) as e:
                logger.error(f"Invalid offer time format: {str(e)}")
                return jsonify({"error": f"Invalid offer time format: {str(e)}"}), 400
        item_data = sanitize_image_fields(item_data)
        item_data['modified_at'] = datetime.now(UTC).isoformat()
        result = items_collection.update_one({'_id': object_id}, {'$set': item_data})
        if result.matched_count == 0:
            logger.warning(f"Item not found for update: {item_id}")
            return jsonify({"error": "Item not found"}), 404
        logger.info(f"Item updated: {item_id}")
        return jsonify({"message": "Item updated successfully"}), 200
    except Exception as e:
        logger.error(f"Error updating item {item_id}: {str(e)}")
        return jsonify({"error": f"Server error: {str(e)}"}), 500

@app.route('/api/items/<item_id>', methods=['PATCH'])
def patch_item(item_id):
    try:
        item_data = request.get_json()
        if not item_data:
            logger.error("No data provided for item patch")
            return jsonify({"error": "No data provided"}), 400
        try:
            object_id = ObjectId(item_id)
        except Exception:
            logger.error(f"Invalid item ID: {item_id}")
            return jsonify({"error": "Invalid item ID"}), 400
        if '_id' in item_data:
            del item_data['_id']
        item_data = sanitize_image_fields(item_data)
        item_data['modified_at'] = datetime.now(UTC).isoformat()
        result = items_collection.update_one({'_id': object_id}, {'$set': item_data})
        if result.matched_count == 0:
            logger.warning(f"Item not found for patch: {item_id}")
            return jsonify({"error": "Item not found"}), 404
        logger.info(f"Item patched: {item_id}")
        return jsonify({"message": "Item updated successfully"}), 200
    except Exception as e:
        logger.error(f"Error patching item {item_id}: {str(e)}")
        return jsonify({"error": f"Server error: {str(e)}"}), 500

@app.route('/api/items/<item_id>', methods=['DELETE'])
def delete_item(item_id):
    try:
        try:
            object_id = ObjectId(item_id)
        except Exception:
            logger.error(f"Invalid item ID: {item_id}")
            return jsonify({"error": "Invalid item ID"}), 400
        result = items_collection.delete_one({'_id': object_id})
        if result.deleted_count == 0:
            logger.warning(f"Item not found for deletion: {item_id}")
            return jsonify({"error": "Item not found"}), 404
        logger.info(f"Item deleted: {item_id}")
        return jsonify({"message": "Item deleted successfully"}), 200
    except Exception as e:
        logger.error(f"Error deleting item {item_id}: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/items/<item_id>/offer', methods=['PUT'])
def update_item_offer(item_id):
    try:
        offer_data = request.get_json()
        try:
            object_id = ObjectId(item_id)
        except Exception:
            logger.error(f"Invalid item ID: {item_id}")
            return jsonify({"error": "Invalid item ID"}), 400
        if 'offer_price' not in offer_data or 'offer_start_time' not in offer_data or 'offer_end_time' not in offer_data:
            return jsonify({"error": "Offer price, start time, and end time are required"}), 400
        try:
            offer_start_time = datetime.fromisoformat(str(offer_data['offer_start_time']).replace('Z', '+00:00'))
            offer_end_time = datetime.fromisoformat(str(offer_data['offer_end_time']).replace('Z', '+00:00'))
            if offer_start_time >= offer_end_time:
                logger.error("offer_start_time must be before offer_end_time")
                return jsonify({"error": "Offer start time must be before offer end time"}), 400
        except (ValueError, TypeError) as e:
            logger.error(f"Invalid offer time format: {str(e)}")
            return jsonify({"error": f"Invalid offer time format: {str(e)}"}), 400
        offer_data['modified_at'] = datetime.now(UTC).isoformat()
        result = items_collection.update_one({'_id': object_id}, {'$set': offer_data})
        if result.matched_count == 0:
            logger.warning(f"Item not found for offer update: {item_id}")
            return jsonify({"error": "Item not found"}), 404
        logger.info(f"Offer updated for item: {item_id}, start: {offer_start_time}, end: {offer_end_time}")
        return jsonify({"message": "Offer updated successfully"}), 200
    except Exception as e:
        logger.error(f"Error updating offer for item {item_id}: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/customers', methods=['GET'])
def get_all_customers():
    try:
        customers = list(customers_collection.find())
        customers = [convert_objectid_to_str(customer) for customer in customers]
        logger.info(f"Fetched {len(customers)} customers")
        return jsonify(customers), 200
    except Exception as e:
        logger.error(f"Error fetching customers: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/customers/<customer_id>', methods=['GET', 'PUT', 'DELETE'])
def customer_operations(customer_id):
    try:
        if not customer_id or customer_id == "undefined":
            logger.error("Invalid customer_id: 'undefined' or empty")
            return jsonify({"error": "Invalid customer ID"}), 400
        try:
            object_id = ObjectId(customer_id)
        except Exception as e:
            logger.error(f"Invalid ObjectId: {customer_id} - {str(e)}")
            return jsonify({"error": "Invalid customer ID format"}), 400
        if request.method == 'GET':
            customer = customers_collection.find_one({'_id': object_id})
            if not customer:
                logger.warning(f"Customer not found: {customer_id}")
                return jsonify({"error": "Customer not found"}), 404
            customer = convert_objectid_to_str(customer)
            logger.info(f"Fetched customer: {customer_id}")
            return jsonify(customer), 200
        elif request.method == 'PUT':
            customer_data = request.get_json()
            if not customer_data:
                logger.error("No data provided for customer update")
                return jsonify({"error": "No data provided"}), 400
            result = customers_collection.update_one(
                {'_id': object_id},
                {'$set': {
                    'customer_name': customer_data.get('customer_name', ''),
                    'phone_number': customer_data.get('phone_number', ''),
                    'whatsapp_number': customer_data.get('whatsapp_number', ''),
                    'email': customer_data.get('email', ''),
                    'building_name': customer_data.get('building_name', ''),
                    'flat_villa_no': customer_data.get('flat_villa_no', ''),
                    'location': customer_data.get('location', ''),
                    'modified_at': datetime.now(UTC).isoformat()
                }}
            )
            if result.matched_count == 0:
                logger.warning(f"Customer not found for update: {customer_id}")
                return jsonify({"error": "Customer not found"}), 404
            logger.info(f"Customer updated: {customer_id}")
            return jsonify({"message": "Customer updated successfully"}), 200
        elif request.method == 'DELETE':
            result = customers_collection.delete_one({'_id': object_id})
            if result.deleted_count == 0:
                logger.warning(f"Customer not found for deletion: {customer_id}")
                return jsonify({"error": "Customer not found"}), 404
            logger.info(f"Customer deleted: {customer_id}")
            return jsonify({"message": "Customer deleted successfully"}), 200
    except Exception as e:
        logger.error(f"Error in customer operations for {customer_id}: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/customers', methods=['POST'])
def create_customer():
    try:
        customer_data = request.get_json()
        if not customer_data or 'customer_name' not in customer_data or 'phone_number' not in customer_data:
            logger.error("Invalid customer data provided")
            return jsonify({"error": "Customer name and phone number are required"}), 400
        existing_customer = customers_collection.find_one({'phone_number': customer_data['phone_number']})
        if existing_customer:
            logger.warning(f"Duplicate phone number: {customer_data['phone_number']}")
            return jsonify({"error": "Phone number already exists", "customer_name": existing_customer['customer_name']}), 409
        customer_data['created_at'] = datetime.now(UTC).isoformat()
        customer_data['modified_at'] = customer_data['created_at']
        result = customers_collection.insert_one(customer_data)
        new_customer_id = str(result.inserted_id)
        logger.info(f"Customer created: {new_customer_id}")
        return jsonify({"id": new_customer_id, "message": "Customer created successfully"}), 201
    except Exception as e:
        logger.error(f"Error creating customer: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/sales', methods=['POST'])
def create_sales_invoice():
    try:
        sales_data = request.json
        required_fields = ['customer', 'items', 'total', 'userId']
        missing_fields = [field for field in required_fields if field not in sales_data or sales_data[field] is None]
        if missing_fields:
            error_msg = f"Missing required fields: {', '.join(missing_fields)}"
            logger.error(error_msg)
            return jsonify({"error": error_msg}), 400
        
        # Validate userId exists in users collection
        user = users_collection.find_one({"email": sales_data['userId']})
        if not user:
            logger.error(f"Invalid userId: {sales_data['userId']}")
            return jsonify({"error": "Invalid userId"}), 400
        
        # Set date and time if not provided
        sales_data['date'] = sales_data.get('date', datetime.now().strftime("%Y-%m-%d"))
        sales_data['time'] = sales_data.get('time', datetime.now().strftime("%H:%M:%S"))
        
        # Calculate totals if not fully provided
        net_total = float(sales_data['total'])
        vat_amount = float(sales_data.get('vat_amount', net_total * 0.10))
        grand_total = float(sales_data.get('grand_total', net_total + vat_amount))
        sales_data['vat_amount'] = round(vat_amount, 2)
        sales_data['grand_total'] = round(grand_total, 2)
        
        # Ensure invoice_no is present
        sales_data['invoice_no'] = sales_data.get('invoice_no', f"INV-{int(datetime.now().timestamp())}")
        sales_data['status'] = sales_data.get('status', 'Draft')
        
        # Process items
        processed_items = []
        for item in sales_data.get('items', []):
            if not all(key in item for key in ['item_name', 'basePrice', 'quantity']):
                logger.error("Invalid item structure in sales invoice")
                return jsonify({"error": "Each item must include item_name, basePrice, and quantity"}), 400
            processed_addons = []
            for addon in item.get('addons', []):
                if not all(key in addon for key in ['name1', 'addon_price', 'addon_quantity']):
                    logger.error("Invalid addon structure in sales invoice")
                    return jsonify({"error": "Each addon must include name1, addon_price, and addon_quantity"}), 400
                processed_addons.append({
                    "addon_name": addon['name1'],
                    "addon_price": float(addon['addon_price']),
                    "addon_quantity": int(addon['addon_quantity']),
                    "addon_image": addon.get('addon_image', ''),
                    "size": addon.get('size', 'M'),
                    "kitchen": addon.get('kitchen', 'Main Kitchen'),
                })
            processed_combos = []
            for combo in item.get('selectedCombos', []):
                if not all(key in combo for key in ['name1', 'combo_price']):
                    logger.error("Invalid combo structure in sales invoice")
                    return jsonify({"error": "Each combo must include name1 and combo_price"}), 400
                processed_combos.append({
                    "name1": combo['name1'],
                    "combo_price": float(combo['combo_price']),
                    "combo_quantity": int(combo.get('combo_quantity', 1)),
                    "combo_image": combo.get('combo_image', ''),
                    "size": combo.get('size', 'M'),
                    "spicy": combo.get('spicy', False),
                    "kitchen": combo.get('kitchen', 'Main Kitchen'),
                })
            processed_items.append({
                "item_name": item['item_name'],
                "basePrice": float(item['basePrice']),
                "quantity": int(item['quantity']),
                "amount": float(item.get('amount', item['basePrice'])),
                "icePreference": item.get('icePreference', 'without_ice'),
                "isSpicy": item.get('isSpicy', False),
                "kitchen": item.get('kitchen', 'Main Kitchen'),
                "selectedSize": item.get('selectedSize', 'M'),
                "ingredients": item.get('ingredients', []),
                "addons": processed_addons,
                "selectedCombos": processed_combos,
            })
        sales_data['items'] = processed_items
        sales_data['created_at'] = datetime.now(UTC).isoformat()
        
        # Insert into database
        sales_id = sales_collection.insert_one(sales_data).inserted_id
        logger.info(f"Sale saved successfully: {sales_data['invoice_no']} by user {sales_data['userId']}")
        
        return jsonify({
            "id": str(sales_id),
            "invoice_no": sales_data['invoice_no'],
            "net_total": sales_data['total'],
            "vat_amount": sales_data['vat_amount'],
            "grand_total": sales_data['grand_total'],
            "userId": sales_data['userId']
        }), 201
    except Exception as e:
        logger.error(f"Error creating sales invoice: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/sales', methods=['GET'])
def get_all_sales():
    try:
        sales = list(sales_collection.find())
        sales = [convert_objectid_to_str(sale) for sale in sales]
        logger.info(f"Fetched {len(sales)} sales invoices")
        return jsonify(sales), 200
    except Exception as e:
        logger.error(f"Error fetching sales: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/sales/<invoice_no>', methods=['GET'])
def get_sale_by_invoice_no(invoice_no):
    try:
        sale = sales_collection.find_one({'invoice_no': invoice_no.strip()})
        if not sale:
            logger.warning(f"Sale not found: {invoice_no}")
            return jsonify({"error": "Invoice not found"}), 404
        sale = convert_objectid_to_str(sale)
        logger.info(f"Fetched sale: {invoice_no}")
        return jsonify(sale), 200
    except Exception as e:
        logger.error(f"Error fetching sale {invoice_no}: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/sales/<invoice_no>/status', methods=['PUT'])
def update_sale_status(invoice_no):
    try:
        data = request.get_json()
        status = data.get('status')
        if not status:
            return jsonify({"error": "Status is required"}), 400
        result = sales_collection.update_one(
            {'invoice_no': invoice_no.strip()},
            {'$set': {'status': status, 'modified_at': datetime.now(UTC).isoformat()}}
        )
        if result.matched_count == 0:
            logger.warning(f"Sale not found for status update: {invoice_no}")
            return jsonify({"error": "Invoice not found"}), 404
        logger.info(f"Sale status updated: {invoice_no} to {status}")
        return jsonify({"message": "Sale status updated successfully"}), 200
    except Exception as e:
        logger.error(f"Error updating sale status {invoice_no}: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/tables', methods=['GET'])
def get_tables():
    try:
        tables = list(tables_collection.find())
        tables = [convert_objectid_to_str(table) for table in tables]
        logger.info(f"Fetched {len(tables)} tables")
        return jsonify({"message": tables}), 200
    except Exception as e:
        logger.error(f"Error fetching tables: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/tables', methods=['POST'])
def add_table():
    try:
        data = request.get_json()
        table_number = data.get("table_number")
        number_of_chairs = data.get("number_of_chairs")
        if not table_number or not number_of_chairs:
            logger.error("Missing table_number or number_of_chairs")
            return jsonify({"error": "Table number and number of chairs are required"}), 400
        if tables_collection.find_one({"table_number": table_number}):
            logger.warning(f"Table number already exists: {table_number}")
            return jsonify({"error": "Table number already exists"}), 400
        new_table = {
            "table_number": table_number,
            "number_of_chairs": int(number_of_chairs),
            "created_at": datetime.now(UTC).isoformat()
        }
        tables_collection.insert_one(new_table)
        logger.info(f"Table added: {table_number}")
        return jsonify({"message": "Table added successfully"}), 201
    except Exception as e:
        logger.error(f"Error adding table: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/tables/<table_number>', methods=['PUT'])
def update_table(table_number):
    try:
        data = request.get_json()
        number_of_chairs = data.get("number_of_chairs")
        if not number_of_chairs:
            return jsonify({"error": "Number of chairs is required"}), 400
        result = tables_collection.update_one(
            {"table_number": table_number},
            {"$set": {"number_of_chairs": int(number_of_chairs), "modified_at": datetime.now(UTC).isoformat()}}
        )
        if result.matched_count == 0:
            logger.warning(f"Table not found for update: {table_number}")
            return jsonify({"error": "Table not found"}), 404
        logger.info(f"Table updated: {table_number}")
        return jsonify({"message": "Table updated successfully"}), 200
    except Exception as e:
        logger.error(f"Error updating table {table_number}: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/tables/<table_number>', methods=['DELETE'])
def delete_table(table_number):
    try:
        result = tables_collection.delete_one({"table_number": table_number})
        if result.deleted_count == 0:
            logger.warning(f"Table not found: {table_number}")
            return jsonify({"error": "Table not found"}), 404
        logger.info(f"Table deleted: {table_number}")
        return jsonify({"message": "Table deleted successfully"}), 200
    except Exception as e:
        logger.error(f"Error deleting table {table_number}: {str(e)}")
        return jsonify({"error": str(e)}), 500







@app.route('/api/create_opening_entry', methods=['POST'])
def create_opening_entry():
    """Create a POS opening entry."""
    try:
        data = request.get_json()
        if not data:
            return jsonify({"message": "No data provided", "status": "error"}), 400
        required_fields = ['period_start_date', 'posting_date', 'company', 'user', 'balance_details']
        missing_fields = [field for field in required_fields if field not in data or not data[field]]
        if missing_fields:
            return jsonify({"message": f"Missing fields: {', '.join(missing_fields)}", "status": "error"}), 400
        balance_details = data['balance_details']
        if not isinstance(balance_details, list):
            return jsonify({"message": "balance_details must be a list", "status": "error"}), 400
        for detail in balance_details:
            if not all(key in detail for key in ['mode_of_payment', 'opening_amount']):
                return jsonify({"message": "Each balance detail must have mode_of_payment and opening_amount", "status": "error"}), 400
            detail['opening_amount'] = float(detail['opening_amount'])
        data['creation'] = datetime.now(UTC).isoformat()
        data['modified'] = data['creation']
        data['name'] = f"OPEN-{int(datetime.now().timestamp())}"
        data['status'] = data.get('status', 'Draft')
        data['docstatus'] = data.get('docstatus', 0)
        data['pos_profile'] = data.get('pos_profile', 'POS-001')
        result = opening_collection.insert_one(data)

        # Update the user's last_opening_entry_time
        username = data['user']
        user = users_collection.find_one({"firstName": username})
        if user:
            users_collection.update_one(
                {"_id": user['_id']},
                {"$set": {"last_opening_entry_time": datetime.now(UTC).isoformat()}}
            )
            logger.info(f"Updated last_opening_entry_time for user: {username}")

        logger.info(f"POS opening entry created: {data['name']}")
        return jsonify({"message": {"name": data['name'], "status": "success"}}), 201
    except Exception as e:
        logger.error(f"Error in create_opening_entry: {str(e)}")
        return jsonify({"message": f"Server error: {str(e)}", "status": "error"}), 500

@app.route('/api/get_pos_opening_entries', methods=['POST'])
def get_pos_opening_entries():
    """Fetch POS opening entries by profile."""
    try:
        data = request.get_json()
        if not data or 'pos_profile' not in data:
            return jsonify({"message": "POS profile is required", "status": "error"}), 400
        pos_profile = data['pos_profile']
        entries = list(opening_collection.find({"pos_profile": pos_profile}))
        entries = [convert_objectid_to_str(entry) for entry in entries]
        logger.info(f"Fetched {len(entries)} POS opening entries for profile: {pos_profile}")
        return jsonify({"message": entries, "status": "success"}), 200
    except Exception as e:
        logger.error(f"Error in get_pos_opening_entries: {str(e)}")
        return jsonify({"message": f"Server error: {str(e)}", "status": "error"}), 500

@app.route('/api/get_pos_invoices', methods=['POST'])
def get_pos_invoices():
    """Fetch POS invoices for a given opening entry."""
    try:
        data = request.get_json()
        pos_opening_entry = data.get('pos_opening_entry')
        if not pos_opening_entry:
            return jsonify({"message": "POS opening entry is required", "status": "error"}), 400
        opening_entry = opening_collection.find_one({"name": pos_opening_entry})
        if not opening_entry:
            return jsonify({"message": "Opening entry not found", "status": "error"}), 404
        period_start = opening_entry['period_start_date']
        invoices = list(sales_collection.find({"date": {"$gte": period_start}}))
        invoices = [convert_objectid_to_str(inv) for inv in invoices]
        total = sum(float(inv['grand_total']) for inv in invoices)
        net_total = sum(float(inv['total']) for inv in invoices)
        total_qty = sum(sum(item['quantity'] for item in inv['items']) for inv in invoices)
        taxes = [{"account_head": "VAT", "rate": 10, "amount": total - net_total}]
        response = {
            "invoices": [{"pos_invoice": inv['invoice_no'], "grand_total": inv['grand_total'], "posting_date": inv['date'], "customer": inv['customer']} for inv in invoices],
            "taxes": taxes,
            "grand_total": total,
            "net_total": net_total,
            "total_quantity": total_qty,
            "status": "success"
        }
        logger.info(f"Fetched POS invoices for opening entry: {pos_opening_entry}")
        return jsonify({"message": response}), 200
    except Exception as e:
        logger.error(f"Error in get_pos_invoices: {str(e)}")
        return jsonify({"message": f"Error: {str(e)}", "status": "error"}), 500

@app.route('/api/create_closing_entry', methods=['POST'])
def create_closing_entry():
    """Create a POS closing entry."""
    try:
        data = request.get_json()
        required_fields = ['pos_opening_entry', 'posting_date', 'period_end_date', 'pos_transactions', 'payment_reconciliation', 'taxes', 'grand_total', 'net_total', 'total_quantity']
        if not data or not all(field in data for field in required_fields):
            return jsonify({"message": f"Missing fields: {', '.join([f for f in required_fields if f not in data])}", "status": "error"}), 400
        opening_entry = opening_collection.find_one({"name": data['pos_opening_entry']})
        if not opening_entry:
            return jsonify({"message": "Opening entry not found", "status": "error"}), 404
        data['creation'] = datetime.now(UTC).isoformat()
        data['modified'] = data['creation']
        data['name'] = f"CLOSE-{int(datetime.now().timestamp())}"
        data['status'] = 'Draft'
        data['docstatus'] = 0
        result = pos_closing_collection.insert_one(data)
        logger.info(f"POS closing entry created: {data['name']}")
        return jsonify({"message": {"name": data['name'], "status": "success", "message": "Closing Entry created"}}), 201
    except Exception as e:
        logger.error(f"Error in create_closing_entry: {str(e)}")
        return jsonify({"message": f"Error: {str(e)}", "status": "error"}), 500

# @app.route('/api/send-email', methods=['POST', 'OPTIONS'])
# def send_email():
#     """Send an email with HTML content."""
#     if request.method == 'OPTIONS':
#         response = jsonify({"success": True})
#         response.headers['Access-Control-Allow-Origin'] = '*'
#         response.headers['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
#         response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
#         return response, 200
#     try:
#         data = request.get_json()
#         logger.info(f"Received email request: {data}")
#         if not data:
#             logger.error("No data received in send-email request")
#             return jsonify({"success": False, "message": "No data provided"}), 400
#         to_email = data.get('to')
#         subject = data.get('subject')
#         html_content = data.get('html')
#         if not all([to_email, subject, html_content]):
#             logger.error("Missing required email fields")
#             return jsonify({"success": False, "message": "Missing required fields: to, subject, html"}), 400
#         msg = MIMEMultipart('alternative')
#         msg['From'] = FROM_EMAIL
#         msg['To'] = to_email
#         msg['Subject'] = subject
#         msg.attach(MIMEText(html_content, 'html'))
#         with smtplib.SMTP('smtp.gmail.com', 587) as server:
#             server.starttls()
#             server.login(EMAIL_USER, EMAIL_PASS)
#             server.send_message(msg)
#             logger.info(f"Email sent successfully to {to_email}")
#         return jsonify({"success": True, "message": "Email sent successfully"}), 200
#     except smtplib.SMTPAuthenticationError as e:
#         logger.error(f"SMTP Authentication Error: {str(e)}")
#         return jsonify({"success": False, "message": "Email authentication failed"}), 401
#     except smtplib.SMTPException as e:
#         logger.error(f"SMTP Error: {str(e)}")
#         return jsonify({"success": False, "message": f"SMTP error: {str(e)}"}), 500
#     except Exception as e:
#         logger.error(f"Unexpected error sending email: {str(e)}")
#         return jsonify({"success": False, "message": f"Failed to send email: {str(e)}"}), 500

# @app.route('/api/export-all-to-excel', methods=['GET'])
# def export_all_to_excel():
#     """Export all data to an Excel file."""
#     try:
#         wb = openpyxl.Workbook()
#         wb.remove(wb.active)
#         collections = {
#             'customers': customers_collection,
#             'items': items_collection,
#             'sales': sales_collection,
#             'tables': tables_collection,
#             'users': users_collection,
#             'picked_up_items': picked_up_collection,
#             'pos_opening_entries': opening_collection,
#             'pos_closing_entries': pos_closing_collection,
#             'system_settings': settings_collection,
#             'kitchens': kitchens_collection,
#             'item_groups': item_groups_collection
#         }
#         for collection_name, collection in collections.items():
#             ws = wb.create_sheet(title=collection_name)
#             data = list(collection.find())
#             if not data:
#                 ws.append(['No data'])
#                 continue
#             sample_doc = data[0]
#             headers = list(sample_doc.keys())
#             ws.append(headers)
#             for doc in data:
#                 row = [str(doc.get(header, '')) if isinstance(doc.get(header), (ObjectId, list, dict)) else doc.get(header, '') for header in headers]
#                 ws.append(row)
#         buffer = BytesIO()
#         wb.save(buffer)
#         buffer.seek(0)
#         filename = f'restaurant_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
#         logger.info(f"Exported data to Excel: {filename}")
#         return Response(
#             buffer.getvalue(),
#             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
#             headers={'Content-Disposition': f'attachment; filename={filename}'}
#         )
#     except Exception as e:
#         logger.error(f"Error exporting to Excel: {str(e)}")
#         return jsonify({"error": f"Server error: {str(e)}"}), 500

# @app.route('/api/backup-to-excel', methods=['GET'])
# def backup_to_excel():
#     """Create a backup and serve it as a download."""
#     try:
#         success, message = create_backup()
#         if not success:
#             return jsonify({"error": message}), 500
#         filename = message.split(': ')[1]
#         file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
#         with open(file_path, 'rb') as f:
#             file_data = f.read()
#         return Response(
#             file_data,
#             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
#             headers={'Content-Disposition': f'attachment; filename={filename}'}
#         )
#     except Exception as e:
#         logger.error(f"Error serving backup file {filename}: {str(e)}")
#         return jsonify({"error": f"Server error: {str(e)}"}), 500

@app.route('/api/kitchens', methods=['GET'])
def get_kitchens():
    """Fetch all kitchens."""
    try:
        kitchens = list(kitchens_collection.find())
        kitchens = [convert_objectid_to_str(kitchen) for kitchen in kitchens]
        logger.info(f"Fetched {len(kitchens)} kitchens")
        return jsonify(kitchens), 200
    except Exception as e:
        logger.error(f"Error fetching kitchens: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/kitchens', methods=['POST'])
def create_kitchen():
    """Create a new kitchen."""
    try:
        data = request.get_json()
        if not data or 'kitchen_name' not in data:
            logger.error("Missing kitchen_name in request")
            return jsonify({"error": "Kitchen name is required"}), 400
        kitchen_name = data['kitchen_name']
        if kitchens_collection.find_one({"kitchen_name": kitchen_name}):
            logger.warning(f"Kitchen already exists: {kitchen_name}")
            return jsonify({"error": "Kitchen name already exists"}), 400
        new_kitchen = {
            "kitchen_name": kitchen_name,
            "created_at": datetime.now(UTC).isoformat()
        }
        result = kitchens_collection.insert_one(new_kitchen)
        logger.info(f"Kitchen created: {kitchen_name}")
        return jsonify({"message": "Kitchen created successfully", "id": str(result.inserted_id)}), 201
    except Exception as e:
        logger.error(f"Error creating kitchen: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/kitchens/<kitchen_id>', methods=['PUT'])
def update_kitchen(kitchen_id):
    """Update an existing kitchen."""
    try:
        data = request.get_json()
        if not data or 'kitchen_name' not in data:
            logger.error("Missing kitchen_name in request")
            return jsonify({"error": "Kitchen name is required"}), 400
        try:
            object_id = ObjectId(kitchen_id)
        except Exception:
            logger.error(f"Invalid kitchen ID: {kitchen_id}")
            return jsonify({"error": "Invalid kitchen ID"}), 400
        result = kitchens_collection.update_one(
            {'_id': object_id},
            {'$set': {'kitchen_name': data['kitchen_name'], 'modified_at': datetime.now(UTC).isoformat()}}
        )
        if result.matched_count == 0:
            logger.warning(f"Kitchen not found for update: {kitchen_id}")
            return jsonify({"error": "Kitchen not found"}), 404
        logger.info(f"Kitchen updated: {kitchen_id}")
        return jsonify({"message": "Kitchen updated successfully"}), 200
    except Exception as e:
        logger.error(f"Error updating kitchen {kitchen_id}: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/kitchens/<kitchen_id>', methods=['DELETE'])
def delete_kitchen(kitchen_id):
    """Delete a kitchen."""
    try:
        try:
            object_id = ObjectId(kitchen_id)
        except Exception:
            logger.error(f"Invalid kitchen ID: {kitchen_id}")
            return jsonify({"error": "Invalid kitchen ID"}), 400
        result = kitchens_collection.delete_one({'_id': object_id})
        if result.deleted_count == 0:
            logger.warning(f"Kitchen not found for deletion: {kitchen_id}")
            return jsonify({"error": "Kitchen not found"}), 404
        logger.info(f"Kitchen deleted: {kitchen_id}")
        return jsonify({"message": "Kitchen deleted successfully"}), 200
    except Exception as e:
        logger.error(f"Error deleting kitchen {kitchen_id}: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/item-groups', methods=['GET'])
def get_item_groups():
    """Fetch all item groups."""
    try:
        item_groups = list(item_groups_collection.find())
        item_groups = [convert_objectid_to_str(group) for group in item_groups]
        logger.info(f"Fetched {len(item_groups)} item groups")
        return jsonify(item_groups), 200
    except Exception as e:
        logger.error(f"Error fetching item groups: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/item-groups', methods=['POST'])
def create_item_group():
    """Create a new item group."""
    try:
        data = request.get_json()
        if not data or 'group_name' not in data:
            logger.error("Missing group_name in request")
            return jsonify({"error": "Group name is required"}), 400
        group_name = data['group_name']
        if item_groups_collection.find_one({"group_name": group_name}):
            logger.warning(f"Item group already exists: {group_name}")
            return jsonify({"error": "Item group name already exists"}), 400
        new_group = {
            "group_name": group_name,
            "created_at": datetime.now(UTC).isoformat()
        }
        result = item_groups_collection.insert_one(new_group)
        logger.info(f"Item group created: {group_name}")
        return jsonify({"message": "Item group created successfully", "id": str(result.inserted_id)}), 201
    except Exception as e:
        logger.error(f"Error creating item group: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/item-groups/<group_id>', methods=['PUT'])
def update_item_group(group_id):
    """Update an existing item group."""
    try:
        data = request.get_json()
        if not data or 'group_name' not in data:
            logger.error("Missing group_name in request")
            return jsonify({"error": "Group name is required"}), 400
        try:
            object_id = ObjectId(group_id)
        except Exception:
            logger.error(f"Invalid group ID: {group_id}")
            return jsonify({"error": "Invalid group ID"}), 400
        result = item_groups_collection.update_one(
            {'_id': object_id},
            {'$set': {'group_name': data['group_name'], 'modified_at': datetime.now(UTC).isoformat()}}
        )
        if result.matched_count == 0:
            logger.warning(f"Item group not found for update: {group_id}")
            return jsonify({"error": "Item group not found"}), 404
        logger.info(f"Item group updated: {group_id}")
        return jsonify({"message": "Item group updated successfully"}), 200
    except Exception as e:
        logger.error(f"Error updating item group {group_id}: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/item-groups/<group_id>', methods=['DELETE'])
def delete_item_group(group_id):
    """Delete an item group."""
    try:
        try:
            object_id = ObjectId(group_id)
        except Exception:
            logger.error(f"Invalid group ID: {group_id}")
            return jsonify({"error": "Invalid group ID"}), 400
        result = item_groups_collection.delete_one({'_id': object_id})
        if result.deleted_count == 0:
            logger.warning(f"Item group not found for deletion: {group_id}")
            return jsonify({"error": "Item group not found"}), 404
        logger.info(f"Item group deleted: {group_id}")
        return jsonify({"message": "Item group deleted successfully"}), 200
    except Exception as e:
        logger.error(f"Error deleting item group {group_id}: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/items/nutrition', methods=['POST', 'OPTIONS'])
def save_item_nutrition():
    """Save or update ingredients data for an item, addon, or combo across all instances."""
    if request.method == 'OPTIONS':
        response = jsonify({"success": True})
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        return response, 200

    try:
        data = request.get_json()
        if not data:
            logger.error("No data provided for saving ingredients")
            return jsonify({"error": "No data provided"}), 400

        # Validate required fields
        required_fields = ['item_name', 'type', 'instances', 'ingredients']
        for field in required_fields:
            if field not in data or data[field] is None:
                logger.error(f"Missing or null required field: {field}")
                return jsonify({"error": f"Missing or null required field: {field}"}), 400

        item_name = data['item_name']
        item_type = data['type']
        instances = data['instances']
        ingredients = data['ingredients']

        # Validate type
        if item_type not in ['item', 'addon', 'combo']:
            logger.error(f"Invalid type: {item_type}")
            return jsonify({"error": "Invalid type, must be 'item', 'addon', or 'combo'"}), 400

        # Validate ingredients
        if not isinstance(ingredients, list):
            logger.error("Ingredients must be a list of objects")
            return jsonify({"error": "Ingredients must be a list of objects"}), 400

        for ingredient in ingredients:
            if not isinstance(ingredient, dict) or not all(key in ingredient for key in ['name', 'small', 'medium', 'large', 'weight', 'nutrition']):
                logger.error("Each ingredient must be an object with name, small, medium, large, weight, and nutrition fields")
                return jsonify({"error": "Each ingredient must be an object with required fields"}), 400

        # Filter out ingredients with empty names
        filtered_ingredients = [ing for ing in ingredients if ing['name'].strip()]

        # Update all instances
        updated_count = 0
        for instance in instances:
            item_id = instance['item_id']
            index = instance.get('index')

            # Find the item by item_id
            item = items_collection.find_one({'_id': ObjectId(item_id)})
            if not item:
                logger.warning(f"Item not found: {item_id}")
                continue

            # Prepare update
            update_query = {'_id': ObjectId(item_id)}
            update_data = {
                '$set': {
                    'modified_at': datetime.now().isoformat()
                }
            }

            if item_type == 'item':
                update_data['$set']['ingredients'] = filtered_ingredients
            elif item_type == 'addon':
                if index is None or not isinstance(index, int):
                    logger.error(f"Index is required for addons in item_id: {item_id}")
                    continue
                update_data['$set'][f'addons.{index}.ingredients'] = filtered_ingredients
            elif item_type == 'combo':
                if index is None or not isinstance(index, int):
                    logger.error(f"Index is required for combos in item_id: {item_id}")
                    continue
                update_data['$set'][f'combos.{index}.ingredients'] = filtered_ingredients

            # Perform the update
            result = items_collection.update_one(update_query, update_data)
            if result.matched_count > 0:
                updated_count += 1
            else:
                logger.warning(f"Item not found for ingredients update: {item_id}")

        if updated_count == 0:
            logger.error(f"No items updated for {item_type}: {item_name}")
            return jsonify({"error": "No items updated, please check instance data"}), 400

        logger.info(f"Ingredients updated for {item_type}: {item_name} across {updated_count} instances")
        return jsonify({"message": "Ingredients saved successfully"}), 200

    except Exception as e:
        logger.error(f"Error saving ingredients for {item_name}: {str(e)}")
        return jsonify({"error": f"Server error: {str(e)}"}), 500

@app.route('/api/items/nutrition/<item_name>', methods=['GET', 'DELETE', 'OPTIONS'])
def handle_item_nutrition(item_name):
    """Fetch or delete ingredients and nutrition data for an item, addon, or combo."""
    if request.method == 'OPTIONS':
        response = jsonify({"success": True})
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'GET, DELETE, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type, X-Instances'
        return response, 200

    try:
        item_type = request.args.get('type')
        item_id = request.args.get('item_id')
        index = request.args.get('index')

        if request.method == 'GET':
            if not item_type or not item_id:
                logger.error("Type and item_id are required")
                return jsonify({"error": "Type and item_id are required"}), 400

            if item_type not in ['item', 'addon', 'combo']:
                logger.error(f"Invalid type: {item_type}")
                return jsonify({"error": "Invalid type, must be 'item', 'addon', or 'combo'"}), 400

            # Find the item by item_id
            item = items_collection.find_one({'_id': ObjectId(item_id)})
            if not item:
                logger.warning(f"Item not found: {item_id}")
                return jsonify({"error": "Item not found"}), 404

            ingredients = []
            if item_type == 'item':
                ingredients = item.get('ingredients', [])
            elif item_type == 'addon':
                if index is None or not index.isdigit():
                    logger.error("Index is required for addons")
                    return jsonify({"error": "Index is required for addons"}), 400
                index = int(index)
                if index < len(item.get('addons', [])):
                    ingredients = item['addons'][index].get('ingredients', [])
            elif item_type == 'combo':
                if index is None or not index.isdigit():
                    logger.error("Index is required for combos")
                    return jsonify({"error": "Index is required for combos"}), 400
                index = int(index)
                if index < len(item.get('combos', [])):
                    ingredients = item['combos'][index].get('ingredients', [])

            response_data = {
                'ingredients': ingredients,
                'nutrition': {}  # Kept for backward compatibility, can be removed if not used
            }
            logger.info(f"Fetched nutrition data for {item_type}: {item_name}")
            return jsonify(response_data), 200

        if request.method == 'DELETE':
            instances = request.headers.get('X-Instances')
            if not instances:
                logger.error("Instances header is required for DELETE")
                return jsonify({"error": "Instances header is required"}), 400

            try:
                instances = json.loads(instances)
            except json.JSONDecodeError:
                logger.error("Invalid instances header format")
                return jsonify({"error": "Invalid instances header format"}), 400

            if item_type not in ['item', 'addon', 'combo']:
                logger.error(f"Invalid type: {item_type}")
                return jsonify({"error": "Invalid type, must be 'item', 'addon', or 'combo'"}), 400

            # Update all instances
            deleted_count = 0
            for instance in instances:
                item_id = instance['item_id']
                index = instance.get('index')

                # Find the item by item_id
                item = items_collection.find_one({'_id': ObjectId(item_id)})
                if not item:
                    logger.warning(f"Item not found: {item_id}")
                    continue

                update_query = {'_id': ObjectId(item_id)}
                update_data = {
                    '$unset': {},
                    '$set': {
                        'modified_at': datetime.now().isoformat()
                    }
                }

                if item_type == 'item':
                    update_data['$unset']['ingredients'] = ''
                elif item_type == 'addon':
                    if index is None or not isinstance(index, int):
                        logger.error(f"Index is required for addons in item_id: {item_id}")
                        continue
                    update_data['$unset'][f'addons.{index}.ingredients'] = ''
                elif item_type == 'combo':
                    if index is None or not isinstance(index, int):
                        logger.error(f"Index is required for combos in item_id: {item_id}")
                        continue
                    update_data['$unset'][f'combos.{index}.ingredients'] = ''

                result = items_collection.update_one(update_query, update_data)
                if result.matched_count > 0:
                    deleted_count += 1
                else:
                    logger.warning(f"Item not found for nutrition deletion: {item_id}")

            if deleted_count == 0:
                logger.error(f"No items updated for deletion of {item_type}: {item_name}")
                return jsonify({"error": "No items updated for deletion, please check instance data"}), 400

            logger.info(f"Nutrition and ingredients cleared for {item_type}: {item_name} across {deleted_count} instances")
            return jsonify({"message": "Nutrition and ingredients cleared successfully"}), 200

    except Exception as e:
        logger.error(f"Error handling nutrition for {item_name}: {str(e)}")
        return jsonify({"error": f"Server error: {str(e)}"}), 500




@app.route('/api/kitchen-saved', methods=['POST'])
def save_kitchen_order():
    try:
        data = request.get_json()
        if not data or 'orderId' not in data:
            return jsonify({'success': False, 'error': 'No data or orderId provided'}), 400

        order_id = data['orderId']
        existing_order = kitchen_saved_collection.find_one({'orderId': order_id})
        
        cart_items = data.get('cartItems', [])
        for item in cart_items:
            required_kitchens = set()
            if 'kitchen' in item:
                required_kitchens.add(item['kitchen'])
            for addon_name, qty in item.get('addonQuantities', {}).items():
                if qty > 0 and 'addonVariants' in item and addon_name in item['addonVariants']:
                    addon = item['addonVariants'][addon_name]
                    if 'kitchen' in addon:
                        required_kitchens.add(addon['kitchen'])
            for combo_name, qty in item.get('comboQuantities', {}).items():
                if qty > 0 and 'comboVariants' in item and combo_name in item['comboVariants']:
                    combo = item['comboVariants'][combo_name]
                    if 'kitchen' in combo:
                        required_kitchens.add(combo['kitchen'])
            item['requiredKitchens'] = list(required_kitchens)
            item['kitchenStatuses'] = item.get('kitchenStatuses', {kitchen: 'Pending' for kitchen in required_kitchens})

        order = {
            'orderId': order_id,
            'customerName': data.get('customerName', 'N/A'),
            'tableNumber': data.get('tableNumber', 'N/A'),
            'chairsBooked': data.get('chairsBooked', []),
            'phoneNumber': data.get('phoneNumber', ''),
            'deliveryAddress': data.get('deliveryAddress', {}),
            'whatsappNumber': data.get('whatsappNumber', ''),
            'email': data.get('email', ''),
            'cartItems': cart_items,
            'timestamp': data.get('timestamp', datetime.utcnow().isoformat()),
            'orderType': data.get('orderType', 'Dine In'),
            'createdAt': datetime.utcnow(),
            'status': data.get('status', 'Pending'),
            'pickedUpTime': data.get('pickedUpTime', None)
        }

        if existing_order:
            kitchen_saved_collection.update_one(
                {'orderId': order_id},
                {'$set': order}
            )
            logger.info(f"Updated kitchen order: {order_id}")
        else:
            kitchen_saved_collection.insert_one(order)
            logger.info(f"Created kitchen order: {order_id}")

        return jsonify({'success': True, 'order_id': order_id}), 201
    except Exception as e:
        logger.error(f"Error in /api/kitchen-saved POST: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/kitchen-saved', methods=['GET'])
def get_kitchen_orders():
    try:
        orders = list(kitchen_saved_collection.find({}, {'_id': 0}))
        return jsonify({'success': True, 'orders': orders}), 200
    except Exception as e:
        logger.error(f"Error in /api/kitchen-saved GET: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/kitchen-saved/<order_id>', methods=['DELETE'])
def delete_kitchen_order(order_id):
    try:
        result = kitchen_saved_collection.delete_one({'orderId': order_id})
        if result.deleted_count == 0:
            logger.warning(f"Order not found: {order_id}")
            return jsonify({'success': False, 'error': 'Order not found'}), 404
        logger.info(f"Order deleted: {order_id}")
        return jsonify({'success': True, 'message': 'Order deleted successfully'}), 200
    except Exception as e:
        logger.error(f"Error deleting order {order_id}: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/kitchen-saved/<order_id>/items/<item_id>/mark-prepared', methods=['POST'])
def mark_item_prepared(order_id, item_id):
    try:
        data = request.get_json()
        kitchen = data.get('kitchen')
        if not kitchen:
            return jsonify({'success': False, 'error': 'Kitchen not provided'}), 400

        order = kitchen_saved_collection.find_one({'orderId': order_id})
        if not order:
            return jsonify({'success': False, 'error': 'Order not found'}), 404

        item = next((item for item in order['cartItems'] if item['id'] == item_id), None)
        if not item:
            return jsonify({'success': False, 'error': 'Item not found'}), 404

        if not item.get('requiredKitchens') or kitchen not in item['requiredKitchens']:
            return jsonify({'success': False, 'error': 'Kitchen not required for this item'}), 400

        if not item.get('kitchenStatuses'):
            item['kitchenStatuses'] = {k: 'Pending' for k in item['requiredKitchens']}

        if item['kitchenStatuses'][kitchen] in ['Prepared', 'PickedUp']:
            return jsonify({'success': False, 'error': 'Kitchen already marked as prepared or picked up'}), 400

        item['kitchenStatuses'][kitchen] = 'Prepared'

        kitchen_saved_collection.update_one(
            {'orderId': order_id, 'cartItems.id': item_id},
            {'$set': {'cartItems.$.kitchenStatuses': item['kitchenStatuses']}}
        )
        activeorders_collection.update_one(
            {'orderId': order_id, 'cartItems.id': item_id},
            {'$set': {'cartItems.$.kitchenStatuses': item['kitchenStatuses']}}
        )

        return jsonify({'success': True, 'status': 'Prepared'}), 200
    except Exception as e:
        logger.error(f"Error in /api/kitchen-saved/{order_id}/items/{item_id}/mark-prepared: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500


def convert_objectid_to_str(data):
    if isinstance(data, list):
        return [convert_objectid_to_str(item) for item in data]
    elif isinstance(data, dict):
        return {key: convert_objectid_to_str(value) for key, value in data.items()}
    elif isinstance(data, ObjectId):
        return str(data)
    return data

# Save picked-up item (POST /api/picked-up-items)
@app.route('/api/picked-up-items', methods=['POST'])
def save_picked_up_item():
    try:
        item_data = request.get_json()
        if not item_data:
            logger.error("No data provided for picked-up items")
            return jsonify({'success': False, 'message': 'No data provided'}), 400

        customer_name = item_data.get('customerName', 'Unknown')
        table_number = item_data.get('tableNumber', 'N/A')
        pickup_time = datetime.utcnow().isoformat()

        new_item = {
            'itemName': item_data.get('itemName', 'Unknown'),
            'quantity': item_data.get('quantity', 0),
            'category': item_data.get('category', 'N/A'),
            'kitchen': item_data.get('kitchen', 'Unknown'),
            'addonCounts': item_data.get('addonCounts', []),
            'selectedCombos': item_data.get('selectedCombos', [])
        }

        existing_entry = picked_up_collection.find_one({
            'customerName': customer_name,
            'tableNumber': table_number
        })

        if existing_entry:
            updated_items = existing_entry.get('items', [])
            updated_items.append(new_item)
            result = picked_up_collection.update_one(
                {'_id': existing_entry['_id']},
                {
                    '$set': {
                        'items': updated_items,
                        'pickupTime': pickup_time,
                        'modified_at': datetime.utcnow().isoformat()
                    }
                }
            )
            logger.info(f"Picked-up items updated for customer: {customer_name}, table: {table_number}")
            return jsonify({
                'success': True,
                'message': 'Picked-up items updated successfully',
                'id': str(existing_entry['_id'])
            }), 200
        else:
            picked_up_data = {
                'customerName': customer_name,
                'tableNumber': table_number,
                'items': [new_item],
                'pickupTime': pickup_time,
                'created_at': datetime.utcnow().isoformat(),
                'orderType': item_data.get('orderType', 'N/A')
            }
            result = picked_up_collection.insert_one(picked_up_data)
            logger.info(f"Picked-up items saved with ID: {result.inserted_id}")
            return jsonify({
                'success': True,
                'message': 'Picked-up items saved successfully',
                'id': str(result.inserted_id)
            }), 201

    except Exception as e:
        logger.error(f"Error saving picked-up items: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500

# Fetch all picked-up items (GET /api/picked-up-items)
@app.route('/api/picked-up-items', methods=['GET'])
def get_picked_up_items():
    try:
        picked_up_items = list(picked_up_collection.find({}))
        # Convert ObjectId to string for JSON serialization
        picked_up_items = convert_objectid_to_str(picked_up_items)
        logger.info(f"Fetched {len(picked_up_items)} picked-up item entries")
        return jsonify({'success': True, 'pickedUpItems': picked_up_items}), 200
    except Exception as e:
        logger.error(f"Error fetching picked-up items: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500

# Delete a picked-up item entry (DELETE /api/picked-up-items/<entry_id>)
@app.route('/api/picked-up-items/<entry_id>', methods=['DELETE'])
def delete_picked_up_item(entry_id):
    try:
        result = picked_up_collection.delete_one({'_id': ObjectId(entry_id)})
        if result.deleted_count == 0:
            logger.warning(f"Picked-up entry not found: {entry_id}")
            return jsonify({"error": "Picked-up entry not found"}), 404
        logger.info(f"Picked-up entry deleted: {entry_id}")
        return jsonify({"message": "Picked-up entry deleted successfully"}), 200
    except Exception as e:
        logger.error(f"Error deleting picked-up entry {entry_id}: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({"error": str(e)}), 500



@app.route('/api/variants', methods=['POST'])
def create_variants():
    try:
        # Get the variant data from the request
        data = request.get_json()
        if not data or not data.get('heading') or not isinstance(data.get('subheadings'), list):
            return jsonify({'error': 'Variant must have a heading and a list of subheadings'}), 400

        # Validate data
        for subheading in data['subheadings']:
            if not subheading.get('name'):
                return jsonify({'error': 'Each subheading must have a name'}), 400
            # Validate price if provided
            if 'price' in subheading and subheading['price'] is not None:
                try:
                    subheading['price'] = float(subheading['price'])
                except (ValueError, TypeError):
                    return jsonify({'error': f"Invalid price for subheading {subheading['name']}"}), 400
            # Validate image if provided
            if 'image' in subheading and subheading['image'] is not None:
                if not isinstance(subheading['image'], str):
                    return jsonify({'error': f"Image for subheading {subheading['name']} must be a string"}), 400
            # Validate dropdown if provided
            if 'dropdown' in subheading and not isinstance(subheading['dropdown'], bool):
                return jsonify({'error': f"Dropdown for subheading {subheading['name']} must be a boolean"}), 400

        # Insert variant into MongoDB
        result = variants_collection.insert_one(data)
        return jsonify({
            'message': 'Variant created successfully',
            'inserted_id': str(result.inserted_id)
        }), 201

    except Exception as e:
        return jsonify({'error': f"Server error: {str(e)}"}), 500

# Route to get all variants
@app.route('/api/variants', methods=['GET'])
def get_variants():
    try:
        # Retrieve all variants from MongoDB
        variants = list(variants_collection.find({}, {'_id': 1, 'heading': 1, 'subheadings': 1, 'activeSection': 1}))
        # Convert ObjectId to string
        for variant in variants:
            variant['_id'] = str(variant['_id'])
        return jsonify(variants), 200
    except Exception as e:
        return jsonify({'error': f"Server error: {str(e)}"}), 500

# Route to get a specific variant by ID
@app.route('/api/variants/<id>', methods=['GET'])
def get_variant(id):
    try:
        # Validate ObjectId
        if not ObjectId.is_valid(id):
            return jsonify({'error': 'Invalid variant ID'}), 400
        variant = variants_collection.find_one({'_id': ObjectId(id)}, {'_id': 0, 'heading': 1, 'subheadings': 1, 'activeSection': 1})
        if not variant:
            return jsonify({'error': 'Variant not found'}), 404
        return jsonify(variant), 200
    except Exception as e:
        return jsonify({'error': f"Server error: {str(e)}"}), 500

# Route to update a variant
@app.route('/api/variants/<id>', methods=['PUT'])
def update_variant(id):
    try:
        # Validate ObjectId
        if not ObjectId.is_valid(id):
            return jsonify({'error': 'Invalid variant ID'}), 400
        data = request.get_json()
        if not data or not data.get('heading') or not isinstance(data.get('subheadings'), list):
            return jsonify({'error': 'Variant must have a heading and a list of subheadings'}), 400

        # Validate data
        for subheading in data['subheadings']:
            if not subheading.get('name'):
                return jsonify({'error': 'Each subheading must have a name'}), 400
            if 'price' in subheading and subheading['price'] is not None:
                try:
                    subheading['price'] = float(subheading['price'])
                except (ValueError, TypeError):
                    return jsonify({'error': f"Invalid price for subheading {subheading['name']}"}), 400
            if 'image' in subheading and subheading['image'] is not None:
                if not isinstance(subheading['image'], str):
                    return jsonify({'error': f"Image for subheading {subheading['name']} must be a string"}), 400
            if 'dropdown' in subheading and not isinstance(subheading['dropdown'], bool):
                return jsonify({'error': f"Dropdown for subheading {subheading['name']} must be a boolean"}), 400

        # Update variant in MongoDB
        result = variants_collection.update_one(
            {'_id': ObjectId(id)},
            {'$set': data}
        )
        if result.matched_count == 0:
            return jsonify({'error': 'Variant not found'}), 404
        return jsonify({'message': 'Variant updated successfully'}), 200
    except Exception as e:
        return jsonify({'error': f"Server error: {str(e)}"}), 500

# Route to delete a variant by ID
@app.route('/api/variants/<id>', methods=['DELETE'])
def delete_variant(id):
    try:
        # Validate ObjectId
        if not ObjectId.is_valid(id):
            return jsonify({'error': 'Invalid variant ID'}), 400
        result = variants_collection.delete_one({'_id': ObjectId(id)})
        if result.deleted_count == 0:
            return jsonify({'error': 'Variant not found'}), 404
        return jsonify({'message': 'Variant deleted successfully'}), 200
    except Exception as e:
        return jsonify({'error': f"Server error: {str(e)}"}), 500

# Route to delete a variant by heading
@app.route('/api/variants/heading/<heading>', methods=['DELETE'])
def delete_variant_by_heading(heading):
    try:
        result = variants_collection.delete_one({'heading': heading})
        return jsonify({'message': 'Variant deleted successfully'}), 200
    except Exception as e:
        return jsonify({'error': f"Server error: {str(e)}"}), 500




VALID_COUNTRY_CODES = [
    '+91',  # India
    '+1',   # USA
    '+971', # UAE (Dubai)
    '+44',  # UK
    '+61',  # Australia
    # Add more country codes as needed
]

def generate_employee_id():
    """Generate a unique employee ID."""
    return str(uuid.uuid4())[:8]  # Simple 8-character UUID for employee ID

def validate_email(email):
    """Validate email format."""
    import re
    pattern = r'^[\w\.-]+@[\w\.-]+\.\w+$'
    return re.match(pattern, email) is not None

@app.route('/api/employees', methods=['GET'])
def get_employees():
    try:
        employees = list(employees_collection.find({}, {'_id': 0}))
        logger.info(f"Fetched {len(employees)} employees")
        return jsonify(employees), 200
    except Exception as e:
        logger.error(f"Error fetching employees: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/employees', methods=['POST'])
def create_employee():
    try:
        data = request.get_json()
        required_fields = ['name', 'phoneNumber', 'vehicleNumber', 'role', 'email']
        if not data or not all(key in data for key in required_fields):
            return jsonify({'error': 'Missing required fields'}), 400

        phone_number = data['phoneNumber']
        if not any(phone_number.startswith(code) for code in VALID_COUNTRY_CODES):
            return jsonify({'error': 'Phone number must include a valid country code (e.g., +91, +1, +971)'}), 400

        code_length = len(next(code for code in VALID_COUNTRY_CODES if phone_number.startswith(code)))
        if len(phone_number) < code_length + 7:
            return jsonify({'error': 'Phone number is too short'}), 400

        email = data['email']
        if not validate_email(email):
            return jsonify({'error': 'Invalid email format'}), 400

        # Check if email is unique
        if employees_collection.find_one({'email': email}):
            return jsonify({'error': 'Email already exists'}), 400

        employee_id = generate_employee_id()
        employee = {
            'employeeId': employee_id,
            'name': data['name'],
            'phoneNumber': phone_number,
            'vehicleNumber': data['vehicleNumber'],
            'role': data['role'],
            'email': email
        }

        # Insert employee into employees_collection
        employees_collection.insert_one(employee)

        # Ensure email exists in users_collection for /api/sales compatibility
        if not users_collection.find_one({'email': email}):
            users_collection.insert_one({
                'email': email,
                'name': data['name'],
                'role': data['role'],
                'created_at': datetime.now(UTC).isoformat()
            })

        employee.pop('_id', None)
        logger.info(f"Created employee: {employee_id} with email: {email}")
        return jsonify({'message': 'Employee created successfully', 'employee': employee}), 201
    except Exception as e:
        logger.error(f"Error creating employee: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/employees/<employee_id>', methods=['PUT'])
def update_employee(employee_id):
    try:
        data = request.get_json()
        required_fields = ['name', 'phoneNumber', 'vehicleNumber', 'role', 'email']
        if not data or not all(key in data for key in required_fields):
            return jsonify({'error': 'Missing required fields'}), 400

        phone_number = data['phoneNumber']
        if not any(phone_number.startswith(code) for code in VALID_COUNTRY_CODES):
            return jsonify({'error': 'Phone number must include a valid country code (e.g., +91, +1, +971)'}), 400

        code_length = len(next(code for code in VALID_COUNTRY_CODES if phone_number.startswith(code)))
        if len(phone_number) < code_length + 7:
            return jsonify({'error': 'Phone number is too short'}), 400

        email = data['email']
        if not validate_email(email):
            return jsonify({'error': 'Invalid email format'}), 400

        # Check if email is unique (excluding current employee)
        existing_employee = employees_collection.find_one({'email': email, 'employeeId': {'$ne': employee_id}})
        if existing_employee:
            return jsonify({'error': 'Email already exists'}), 400

        updated_employee = {
            'name': data['name'],
            'phoneNumber': phone_number,
            'vehicleNumber': data['vehicleNumber'],
            'role': data['role'],
            'email': email
        }

        result = employees_collection.update_one(
            {'employeeId': employee_id},
            {'$set': updated_employee}
        )

        if result.matched_count == 0:
            return jsonify({'error': 'Employee not found'}), 404

        # Update or insert user in users_collection
        users_collection.update_one(
            {'email': email},
            {'$set': {
                'email': email,
                'name': data['name'],
                'role': data['role'],
                'updated_at': datetime.now(UTC).isoformat()
            }},
            upsert=True
        )

        logger.info(f"Updated employee: {employee_id} with email: {email}")
        return jsonify({'message': 'Employee updated successfully'}), 200
    except Exception as e:
        logger.error(f"Error updating employee: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/employees/<employee_id>', methods=['DELETE'])
def delete_employee(employee_id):
    try:
        employee = employees_collection.find_one({'employeeId': employee_id}, {'_id': 0})
        if not employee:
            return jsonify({'error': 'Employee not found'}), 404

        result = employees_collection.delete_one({'employeeId': employee_id})
        if result.deleted_count == 0:
            return jsonify({'error': 'Employee not found'}), 404

        # Optionally, remove from users_collection if no other employees use this email
        if not employees_collection.find_one({'email': employee['email']}):
            users_collection.delete_one({'email': employee['email']})

        logger.info(f"Deleted employee: {employee_id}")
        return jsonify({'message': 'Employee deleted successfully'}), 200
    except Exception as e:
        logger.error(f"Error deleting employee: {str(e)}")
        return jsonify({'error': str(e)}), 500
def generate_unique_id():
    return str(uuid.uuid4())

def generate_order_number(order_type):
    """Generate order number based on order type (e.g., T0001, D0001, ON001)."""
    prefix = {'Dine In': 'D', 'Take Away': 'T', 'Online Delivery': 'ON'}.get(order_type, 'D')
    counter = order_counters_collection.find_one_and_update(
        {'order_type': order_type},
        {'$inc': {'counter': 1}},
        upsert=True,
        return_document=True
    )
    number = counter['counter']
    if prefix == 'ON':
        return f'{prefix}{number:03d}'  # e.g., ON001
    return f'{prefix}{number:04d}'  # e.g., D0001, T0001

# Route to fetch a specific active order by orderId
@app.route('/api/activeorders/<order_id>', methods=['GET'])
def get_active_order(order_id):
    try:
        order = activeorders_collection.find_one({'orderId': order_id}, {'_id': 0})
        if order:
            logger.info(f"Fetched order: {order_id}")
            return jsonify(order), 200
        else:
            logger.warning(f"Order not found: {order_id}")
            return jsonify({'error': 'Order not found'}), 404
    except Exception as e:
        logger.error(f"Error fetching order: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

# Route to create a new active order
@app.route('/api/activeorders', methods=['POST'])
def save_active_order():
    try:
        data = request.get_json()
        order_id = generate_unique_id()
        order_type = data.get('orderType', 'Dine In')
        order_no = generate_order_number(order_type)

        cart_items = data.get('cartItems', [])
        for item in cart_items:
            required_kitchens = set()
            if item.get('kitchen'):
                required_kitchens.add(item['kitchen'])
            
            for addon_name, qty in item.get('addonQuantities', {}).items():
                if qty > 0 and 'addonVariants' in item and addon_name in item['addonVariants']:
                    if item['addonVariants'][addon_name].get('kitchen'):
                        required_kitchens.add(item['addonVariants'][addon_name]['kitchen'])
            
            for combo_name, qty in item.get('comboQuantities', {}).items():
                if qty > 0 and 'comboVariants' in item and combo_name in item['comboVariants']:
                    if item['comboVariants'][combo_name].get('kitchen'):
                        required_kitchens.add(item['comboVariants'][combo_name]['kitchen'])

            item['requiredKitchens'] = list(required_kitchens)
            item['kitchenStatuses'] = {kitchen: 'Pending' for kitchen in required_kitchens}

        active_order = {
            'orderId': order_id,
            'orderNo': order_no,
            'customerName': data.get('customerName', 'N/A'),
            'tableNumber': data.get('tableNumber', 'N/A'),
            'chairsBooked': data.get('chairsBooked', []),
            'phoneNumber': data.get('phoneNumber', ''),
            'deliveryAddress': data.get('deliveryAddress', {}),
            'whatsappNumber': data.get('whatsappNumber', ''),
            'email': data.get('email', ''),
            'cartItems': cart_items,
            'timestamp': data.get('timestamp', datetime.utcnow().isoformat()),
            'orderType': order_type,
            'status': data.get('status', 'Pending'),
            'created_at': datetime.utcnow(),
            'deliveryPersonId': data.get('deliveryPersonId', ''),
            'deliveryPersonName': data.get('deliveryPersonName', ''),
            'pickedUpTime': None
        }

        activeorders_collection.insert_one(active_order)
        kitchen_saved_collection.insert_one(active_order.copy())

        logger.info(f"Created order: {order_id} with order number: {order_no}")
        return jsonify({'success': True, 'orderId': order_id, 'orderNo': order_no}), 201
    except Exception as e:
        logger.error(f"Error saving active order: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

# Route to fetch all active orders
@app.route('/api/activeorders', methods=['GET'])
def get_active_orders():
    try:
        orders = list(activeorders_collection.find({}, {'_id': 0}))
        logger.info(f"Fetched {len(orders)} active orders")
        return jsonify(orders), 200
    except Exception as e:
        logger.error(f"Error fetching active orders: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

# Route to mark an item as prepared
@app.route('/api/activeorders/<order_id>/items/<item_id>/mark-prepared', methods=['POST'])
def mark_item_prepared_active(order_id, item_id):
    try:
        data = request.get_json()
        kitchen = data.get('kitchen')
        if not kitchen:
            return jsonify({'success': False, 'error': 'Kitchen not provided'}), 400

        for collection in [activeorders_collection, kitchen_saved_collection]:
            collection.update_one(
                {'orderId': order_id, 'cartItems.id': item_id},
                {'$set': {f'cartItems.$.kitchenStatuses.{kitchen}': 'Prepared'}}
            )
        
        logger.info(f"Marked item {item_id} in order {order_id} as Prepared for kitchen {kitchen}")
        return jsonify({'success': True, 'status': 'Prepared'}), 200
    except Exception as e:
        logger.error(f"Error in mark-prepared: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500

# Route to mark an item as picked up
@app.route('/api/activeorders/<order_id>/items/<item_id>/mark-pickedup', methods=['POST'])
def mark_item_pickedup_active(order_id, item_id):
    try:
        data = request.get_json()
        kitchen = data.get('kitchen')
        if not kitchen:
            return jsonify({'success': False, 'error': 'Kitchen not provided'}), 400

        order = activeorders_collection.find_one({'orderId': order_id}, {'_id': 0})
        if not order:
            return jsonify({'success': False, 'error': 'Order not found'}), 404

        item = next((item for item in order['cartItems'] if item['id'] == item_id), None)
        if not item:
            return jsonify({'success': False, 'error': 'Item not found'}), 404
        
        if item.get('kitchenStatuses', {}).get(kitchen) != 'Prepared':
            return jsonify({'success': False, 'error': 'Item must be prepared before picking up'}), 400

        for collection in [activeorders_collection, kitchen_saved_collection]:
            collection.update_one(
                {'orderId': order_id, 'cartItems.id': item_id},
                {'$set': {f'cartItems.$.kitchenStatuses.{kitchen}': 'PickedUp'}}
            )

        picked_up_data = {
            'customerName': order.get('customerName', 'Unknown'),
            'tableNumber': order.get('tableNumber', 'N/A'),
            'orderType': order.get('orderType', 'Dine In'),
            'pickupTime': datetime.utcnow().isoformat(),
            'items': [{
                'itemName': item.get('name', 'Unknown'),
                'quantity': item.get('quantity', 0),
                'category': item.get('category', 'N/A'),
                'kitchen': kitchen,
                'addonCounts': [
                    {'name': name, 'quantity': qty}
                    for name, qty in item.get('addonQuantities', {}).items()
                    if qty > 0 and item.get('addonVariants', {}).get(name, {}).get('kitchen') == kitchen
                ],
                'selectedCombos': [
                    {'name': name, 'size': item['comboVariants'][name]['size'], 'quantity': qty}
                    for name, qty in item.get('comboQuantities', {}).items()
                    if qty > 0 and item.get('comboVariants', {}).get(name, {}).get('kitchen') == kitchen
                ]
            }]
        }
        picked_up_collection.insert_one(picked_up_data)
        
        logger.info(f"Marked item {item_id} in order {order_id} as PickedUp for kitchen {kitchen}")
        return jsonify({'success': True, 'status': 'PickedUp'}), 200
    except Exception as e:
        logger.error(f"Error in mark-pickedup: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500

# Route to update an active order
@app.route('/api/activeorders/<order_id>', methods=['PUT'])
def update_active_order(order_id):
    try:
        data = request.get_json()
        if '_id' in data:
            del data['_id']

        order_in_db = activeorders_collection.find_one({'orderId': order_id}, {'_id': 0})
        if not order_in_db:
            logger.warning(f"Order not found for update: {order_id}")
            return jsonify({'error': 'Order not found'}), 404

        old_statuses_map = {
            item['id']: item.get('kitchenStatuses', {}) 
            for item in order_in_db.get('cartItems', []) if 'id' in item
        }

        if 'cartItems' in data:
            for item in data['cartItems']:
                required_kitchens = set()
                if item.get('kitchen'):
                    required_kitchens.add(item['kitchen'])
                for addon_name, qty in item.get('addonQuantities', {}).items():
                    if qty > 0 and item.get('addonVariants', {}).get(addon_name, {}).get('kitchen'):
                        required_kitchens.add(item['addonVariants'][addon_name]['kitchen'])
                for combo_name, qty in item.get('comboQuantities', {}).items():
                    if qty > 0 and item.get('comboVariants', {}).get(combo_name, {}).get('kitchen'):
                        required_kitchens.add(item['comboVariants'][combo_name]['kitchen'])
                
                item['requiredKitchens'] = list(required_kitchens)

                item_id = item.get('id')
                old_item_statuses = old_statuses_map.get(item_id, {})
                
                new_kitchen_statuses = {}
                for kitchen in required_kitchens:
                    if kitchen in old_item_statuses:
                        new_kitchen_statuses[kitchen] = old_item_statuses[kitchen]
                    else:
                        new_kitchen_statuses[kitchen] = 'Pending'
                
                item['kitchenStatuses'] = new_kitchen_statuses

        if 'deliveryPersonId' in data and data['deliveryPersonId']:
            employee = employees_collection.find_one({'employeeId': data['deliveryPersonId']}, {'_id': 0})
            if not employee:
                logger.warning(f"Delivery person not found: {data['deliveryPersonId']}")
                return jsonify({'error': 'Delivery person not found'}), 404

            # Save to trip reports before deleting
            trip_report = {
                'tripId': generate_unique_id(),
                'orderId': order_in_db['orderId'],
                'orderNo': order_in_db['orderNo'],
                'customerName': order_in_db.get('customerName', 'N/A'),
                'tableNumber': order_in_db.get('tableNumber', 'N/A'),
                'chairsBooked': order_in_db.get('chairsBooked', []),
                'phoneNumber': order_in_db.get('phoneNumber', ''),
                'deliveryAddress': order_in_db.get('deliveryAddress', {}),
                'whatsappNumber': order_in_db.get('whatsappNumber', ''),
                'email': order_in_db.get('email', ''),
                'cartItems': order_in_db.get('cartItems', []),
                'timestamp': order_in_db.get('timestamp', datetime.utcnow().isoformat()),
                'orderType': order_in_db.get('orderType', 'Dine In'),
                'status': order_in_db.get('status', 'Pending'),
                'deliveryPersonId': data['deliveryPersonId'],
                'deliveryPersonName': data.get('deliveryPersonName', employee.get('name', 'N/A')),
                'pickedUpTime': order_in_db.get('pickedUpTime', None),
                'paymentMethods': order_in_db.get('paymentMethods', []),
                'cardDetails': order_in_db.get('cardDetails', ''),
                'upiDetails': order_in_db.get('upiDetails', ''),
                'created_at': datetime.utcnow()
            }
            tripreports_collection.insert_one(trip_report)
            logger.info(f"Saved trip report for order {order_id} with delivery person {data['deliveryPersonId']}")

            activeorders_collection.delete_one({'orderId': order_id})
            kitchen_saved_collection.delete_one({'orderId': order_id})
            logger.info(f"Deleted order {order_id} from active orders after delivery person assignment")
            return jsonify({'success': True, 'message': 'Delivery person assigned and order moved to trip reports', 'order': order_in_db}), 200

        result = activeorders_collection.update_one({'orderId': order_id}, {'$set': data})
        kitchen_result = kitchen_saved_collection.update_one({'orderId': order_id}, {'$set': data})

        updated_order = activeorders_collection.find_one({'orderId': order_id}, {'_id': 0})
        if result.modified_count > 0 or kitchen_result.modified_count > 0:
            logger.info(f"Updated order: {order_id}")
            return jsonify({'success': True, 'message': 'Order updated', 'order': updated_order}), 200
        
        logger.info(f"No changes made to order: {order_id}")
        return jsonify({'success': True, 'message': 'No changes made', 'order': updated_order}), 200

    except Exception as e:
        logger.error(f"Error updating active order: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

# Route to delete an active order
@app.route('/api/activeorders/<order_id>', methods=['DELETE'])
def delete_order(order_id):
    try:
        result = activeorders_collection.delete_one({'orderId': order_id})
        kitchen_result = kitchen_saved_collection.delete_one({'orderId': order_id})
        if result.deleted_count > 0 or kitchen_result.deleted_count > 0:
            logger.info(f"Deleted order: {order_id}")
            return jsonify({'success': True}), 200
        logger.warning(f"Order not found for deletion: {order_id}")
        return jsonify({'error': 'Order not found'}), 404
    except Exception as e:
        logger.error(f"Error deleting order: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

# Route to fetch trip reports for an employee
@app.route('/api/tripreports/<employee_id>', methods=['GET'])
def get_trip_reports(employee_id):
    try:
        trip_reports = list(tripreports_collection.find({'deliveryPersonId': employee_id}, {'_id': 0}))
        logger.info(f"Fetched {len(trip_reports)} trip reports for employee: {employee_id}")
        return jsonify(trip_reports), 200
    except Exception as e:
        logger.error(f"Error fetching trip reports: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

shutdown_flag = False

# Example route for testing server
@app.route('/api/test', methods=['GET'])
def test():
    return jsonify({"message": "Server is running"}), 200

# Shutdown endpoint
@app.route('/api/shutdown', methods=['POST'])
def shutdown():
    global shutdown_flag
    logger.info("Shutdown requested")
    
    try:
        # For development (Werkzeug server)
        func = request.environ.get('werkzeug.server.shutdown')
        if func:
            func()
            logger.info("Werkzeug server shutdown initiated")
            return jsonify({"message": "Server shutting down"}), 200
        
        # For production (Waitress server)
        shutdown_flag = True
        logger.info("Setting shutdown flag for Waitress")
        
        # Start a thread to exit the process after a short delay
        def exit_process():
            time.sleep(1)  # Reduced delay to 1 second for faster shutdown
            logger.info("Exiting Python process")
            os._exit(0)  # Forcefully exit the process
        
        threading.Thread(target=exit_process, daemon=True).start()
        return jsonify({"message": "Server shutting down"}), 200
    
    except Exception as e:
        logger.error(f"Shutdown error: {str(e)}")
        return jsonify({"message": "Error during shutdown", "error": str(e)}), 500

@app.route('/api/save-email-settings', methods=['POST'])
def save_email_settings():
    """Save email settings to MongoDB."""
    try:
        data = request.get_json()
        email = data.get('email')
        password = data.get('password')
        from_email = data.get('from_email')
        
        if not all([email, password, from_email]):
            return jsonify({"success": False, "error": "Missing required fields: email, password, from_email"}), 400

        # Update or insert email settings
        email_settings_collection.update_one(
            {},
            {
                '$set': {
                    'email': email,
                    'password': password,
                    'from_email': from_email,
                    'updated_at': datetime.now(UTC)
                }
            },
            upsert=True
        )
        logger.info(f"Email settings saved for {email}")
        return jsonify({"success": True, "message": "Email settings saved successfully"}), 200
    except Exception as e:
        logger.error(f"Error saving email settings: {str(e)}")
        return jsonify({"success": False, "error": f"Failed to save email settings: {str(e)}"}), 500

@app.route('/api/get-email-settings', methods=['GET'])
def get_email_settings():
    """Retrieve email settings from MongoDB."""
    try:
        settings = email_settings_collection.find_one({}, {'_id': 0, 'password': 0})
        if not settings:
            return jsonify({"success": False, "error": "No email settings found"}), 404
        return jsonify({"success": True, "email": settings.get('email'), "from_email": settings.get('from_email')}), 200
    except Exception as e:
        logger.error(f"Error retrieving email settings: {str(e)}")
        return jsonify({"success": False, "error": f"Failed to retrieve email settings: {str(e)}"}), 500

@app.route('/api/test-email-settings', methods=['POST'])
def test_email_settings():
    """Test email settings by attempting to authenticate with SMTP."""
    try:
        data = request.get_json()
        email = data.get('email')
        password = data.get('password')
        if not all([email, password]):
            return jsonify({"success": False, "error": "Missing required fields: email, password"}), 400

        # Attempt SMTP login
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(email, password)
        logger.info(f"SMTP authentication successful for {email}")
        return jsonify({"success": True, "message": "Email settings are valid"}), 200
    except smtplib.SMTPAuthenticationError as e:
        logger.error(f"SMTP Authentication Error during test: {str(e)}")
        return jsonify({"success": False, "error": "Invalid email or app password. Please check your credentials and ensure an App Password is used for Gmail."}), 401
    except smtplib.SMTPException as e:
        logger.error(f"SMTP Error during test: {str(e)}")
        return jsonify({"success": False, "error": f"SMTP error: {str(e)}"}), 500
    except Exception as e:
        logger.error(f"Unexpected error testing email settings: {str(e)}")
        return jsonify({"success": False, "error": f"Failed to test email settings: {str(e)}"}), 500

@app.route('/api/send-email', methods=['POST', 'OPTIONS'])
def send_email():
    """Send an email with HTML content."""
    if request.method == 'OPTIONS':
        response = jsonify({"success": True})
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        return response, 200
    try:
        data = request.get_json()
        logger.info(f"Received email request: {data}")
        if not data:
            logger.error("No data received in send-email request")
            return jsonify({"success": False, "message": "No data provided"}), 400
        to_email = data.get('to')
        subject = data.get('subject')
        html_content = data.get('html')
        if not all([to_email, subject, html_content]):
            logger.error("Missing required email fields")
            return jsonify({"success": False, "message": "Missing required fields: to, subject, html"}), 400
        
        # Fetch email settings
        settings = email_settings_collection.find_one()
        if not settings:
            logger.error("No email settings configured")
            return jsonify({"success": False, "message": "Email settings not configured. Please configure in Email Settings."}), 500
        email_user = settings.get('email')
        email_pass = settings.get('password')
        from_email = settings.get('from_email')

        msg = MIMEMultipart('alternative')
        msg['From'] = from_email
        msg['To'] = to_email
        msg['Subject'] = subject
        msg.attach(MIMEText(html_content, 'html'))
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(email_user, email_pass)
            server.send_message(msg)
            logger.info(f"Email sent successfully to {to_email}")
        return jsonify({"success": True, "message": "Email sent successfully"}), 200
    except smtplib.SMTPAuthenticationError as e:
        logger.error(f"SMTP Authentication Error: {str(e)}")
        return jsonify({"success": False, "message": "Invalid email or app password. Please check your Email Settings and ensure an App Password is used for Gmail."}), 401
    except smtplib.SMTPException as e:
        logger.error(f"SMTP Error: {str(e)}")
        return jsonify({"success": False, "message": f"SMTP error: {str(e)}"}), 500
    except Exception as e:
        logger.error(f"Unexpected error sending email: {str(e)}")
        return jsonify({"success": False, "message": f"Failed to send email: {str(e)}"}), 500

@app.route('/api/export-all-to-excel', methods=['GET'])
def export_all_to_excel():
    """Export all data to an Excel file."""
    try:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        collections = {
            'customers': customers_collection,
            'items': items_collection,
            'sales': sales_collection,
            'tables': tables_collection,
            'users': users_collection,
            'picked_up_items': picked_up_collection,
            'pos_opening_entries': opening_collection,
            'pos_closing_entries': pos_closing_collection,
            'system_settings': settings_collection,
            'kitchens': kitchens_collection,
            'item_groups': item_groups_collection
        }
        for collection_name, collection in collections.items():
            ws = wb.create_sheet(title=collection_name)
            data = list(collection.find())
            if not data:
                ws.append(['No data'])
                continue
            sample_doc = data[0]
            headers = list(sample_doc.keys())
            ws.append(headers)
            for doc in data:
                row = [str(doc.get(header, '')) if isinstance(doc.get(header), (ObjectId, list, dict)) else doc.get(header, '') for header in headers]
                ws.append(row)
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f'restaurant_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        logger.info(f"Exported data to Excel: {filename}")
        return Response(
            buffer.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )
    except Exception as e:
        logger.error(f"Error exporting to Excel: {str(e)}")
        return jsonify({"error": f"Server error: {str(e)}"}), 500

def manage_backup_limit():
    """Manage the number of backup files to keep only the latest MAX_BACKUPS."""
    try:
        backup_files = [f for f in os.listdir(app.config['UPLOAD_FOLDER']) if f.endswith('.xlsx')]
        backup_files = sorted(
            backup_files,
            key=lambda x: os.path.getctime(os.path.join(app.config['UPLOAD_FOLDER'], x)),
            reverse=True
        )
        for old_file in backup_files[MAX_BACKUPS:]:
            os.remove(os.path.join(app.config['UPLOAD_FOLDER'], old_file))
            logger.info(f"Deleted old backup: {old_file}")
    except Exception as e:
        logger.error(f"Error managing backup limit: {str(e)}")

def create_backup():
    """Create a backup and send it via email using settings from MongoDB."""
    try:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        collections = {
            'customers': customers_collection,
            'items': items_collection,
            'sales': sales_collection,
            'tables': tables_collection,
            'users': users_collection,
            'picked_up_items': picked_up_collection,
            'pos_opening_entries': opening_collection,
            'pos_closing_entries': pos_closing_collection,
            'system_settings': settings_collection,
            'kitchens': kitchens_collection,
            'item_groups': item_groups_collection
        }
        for collection_name, collection in collections.items():
            ws = wb.create_sheet(title=collection_name)
            data = list(collection.find())
            if not data:
                ws.append(['No data'])
                continue
            sample_doc = data[0]
            headers = list(sample_doc.keys())
            ws.append(headers)
            for doc in data:
                row = [str(doc.get(header, '')) if isinstance(doc.get(header), (ObjectId, list, dict)) else doc.get(header, '') for header in headers]
                ws.append(row)
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f'backup_restaurant_data_{timestamp}.xlsx'
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        with open(file_path, 'wb') as f:
            f.write(buffer.getvalue())
        manage_backup_limit()

        # Fetch email settings
        settings = email_settings_collection.find_one()
        if not settings:
            logger.error("No email settings configured")
            return False, "Email settings not configured. Please configure in Email Settings."
        email_user = settings.get('email')
        email_pass = settings.get('password')
        from_email = settings.get('from_email')

        # Send email with backup attachment
        msg = MIMEMultipart()
        msg['From'] = from_email
        msg['To'] = email_user
        msg['Subject'] = f'Restaurant Data Backup - {timestamp}'
        body = f'Backup of restaurant data generated on {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}.'
        msg.attach(MIMEText(body, 'plain'))
        with open(file_path, 'rb') as f:
            attachment = MIMEBase('application', 'octet-stream')
            attachment.set_payload(f.read())
            encoders.encode_base64(attachment)
            attachment.add_header('Content-Disposition', f'attachment; filename={filename}')
            msg.attach(attachment)
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(email_user, email_pass)
            server.send_message(msg)
        logger.info(f"Backup created and emailed: {filename}")
        return True, f"Backup created successfully: {filename}"
    except smtplib.SMTPAuthenticationError as e:
        logger.error(f"SMTP Authentication Error: {str(e)}")
        return False, f"Invalid email or app password. Please check your Email Settings and ensure an App Password is used for Gmail."
    except smtplib.SMTPException as e:
        logger.error(f"SMTP Error: {str(e)}")
        return False, f"SMTP error: {str(e)}"
    except Exception as e:
        logger.error(f"Error in backup: {str(e)}")
        return False, str(e)

@app.route('/api/backup-to-excel', methods=['GET'])
def backup_to_excel():
    """Create a backup and serve it as a download."""
    try:
        success, message = create_backup()
        if not success:
            return jsonify({"error": message}), 500
        filename = message.split(': ')[1]
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        with open(file_path, 'rb') as f:
            file_data = f.read()
        return Response(
            file_data,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )
    except Exception as e:
        logger.error(f"Error serving backup file: {str(e)}")
        return jsonify({"error": f"Server error: {str(e)}"}), 500

@app.route('/api/backup-info', methods=['GET'])
def backup_info():
    """Retrieve information about existing backups."""
    try:
        backup_files = [f for f in os.listdir(app.config['UPLOAD_FOLDER']) if f.endswith('.xlsx')]
        backups = []
        for filename in backup_files:
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            stat = os.stat(file_path)
            backups.append({
                'filename': filename,
                'date': datetime.fromtimestamp(stat.st_ctime).strftime('%Y-%m-%d %H:%M:%S'),
                'size': f"{stat.st_size / 1024:.2f} KB"
            })
        backups.sort(key=lambda x: x['date'], reverse=True)
        return jsonify(backups)
    except Exception as e:
        logger.error(f"Error retrieving backup info: {str(e)}")
        return jsonify({"error": f"Server error: {str(e)}"}), 500

@app.route('/api/download-backup', methods=['POST'])
def download_backup():
    """Download a specific backup file."""
    try:
        data = request.get_json()
        filename = data.get('filename')
        if not filename:
            return jsonify({"error": "Filename not provided"}), 400
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        if not os.path.exists(file_path):
            return jsonify({"error": "Backup file not found"}), 404
        with open(file_path, 'rb') as f:
            file_data = f.read()
        return Response(
            file_data,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )
    except Exception as e:
        logger.error(f"Error downloading backup {filename}: {str(e)}")
        return jsonify({"error": f"Server error: {str(e)}"}), 500

def manage_offers():
    """Check all items and update offer status based on current time."""
    try:
        current_time = datetime.now(UTC)
        items = items_collection.find({
            '$or': [
                {'offer_start_time': {'$exists': True}},
                {'offer_end_time': {'$exists': True}}
            ]
        })
        for item in items:
            item_id = item['_id']
            offer_start_time = item.get('offer_start_time')
            offer_end_time = item.get('offer_end_time')
            should_unset = False
            if offer_start_time and offer_end_time:
                try:
                    start_time = datetime.fromisoformat(str(offer_start_time).replace('Z', '+00:00'))
                    end_time = datetime.fromisoformat(str(offer_end_time).replace('Z', '+00:00'))
                    if current_time > end_time:
                        should_unset = True
                        logger.info(f"Offer expired for item {item.get('item_name')} (ID: {item_id})")
                    elif start_time > end_time:
                        should_unset = True
                        logger.warning(f"Invalid offer times for item {item_id}: start_time after end_time")
                    else:
                        logger.debug(f"Offer for item {item.get('item_name')} (ID: {item_id}) is active or pending")
                except (ValueError, TypeError) as e:
                    logger.warning(f"Invalid offer time format for item {item_id}: {str(e)}")
                    should_unset = True
            elif offer_end_time:
                try:
                    end_time = datetime.fromisoformat(str(offer_end_time).replace('Z', '+00:00'))
                    if current_time > end_time:
                        should_unset = True
                        logger.info(f"Offer expired for item {item.get('item_name')} (ID: {item_id})")
                except (ValueError, TypeError) as e:
                    logger.warning(f"Invalid offer_end_time for item {item_id}: {str(e)}")
                    should_unset = True
            if should_unset:
                items_collection.update_one(
                    {'_id': item_id},
                    {'$unset': {'offer_price': "", 'offer_start_time': "", 'offer_end_time': ""}}
                )
                logger.info(f"Unset offer fields for item {item.get('item_name')} (ID: {item_id})")
    except Exception as e:
        logger.error(f"Error in manage_offers: {str(e)}")

def schedule_tasks():
    """Schedule backups and offer management."""
    schedule.every(6).hours.do(create_backup)
    schedule.every(1).minutes.do(manage_offers)
    while True:
        schedule.run_pending()
        time.sleep(60)

def start_scheduler():
    scheduler_thread = threading.Thread(target=schedule_tasks, daemon=True)
    scheduler_thread.start()
    logger.info("Automatic backup and offer scheduler started")


def document_to_dict(doc):
    doc['_id'] = str(doc['_id'])
    if 'created_at' in doc:
        doc['created_at'] = doc['created_at'].isoformat()
    if 'date' in doc:
        doc['date'] = doc['date'].isoformat()
    return doc

# --- Purchase Items Routes ---
@app.route('/api/purchase_items', methods=['GET'])
def get_purchase_items():
    try:
        items = list(purchase_items_collection.find())
        return jsonify([document_to_dict(item) for item in items]), 200
    except Exception as e:
        return jsonify({'error': f"Failed to fetch items: {str(e)}"}), 500

@app.route('/api/purchase_items', methods=['POST'])
def add_purchase_item():
    try:
        data = request.json
        required_fields = ['name', 'mainUnit', 'subUnit', 'conversionFactor']
        if not all(key in data for key in required_fields):
            return jsonify({'error': 'Missing required fields'}), 400
        if not data['name'] or not data['mainUnit'] or not data['subUnit'] or float(data['conversionFactor']) <= 0:
            return jsonify({'error': 'Invalid input data'}), 400
        
        item = {
            'name': data['name'],
            'mainUnit': data['mainUnit'],
            'subUnit': data['subUnit'],
            'conversionFactor': float(data['conversionFactor']),
            'created_at': datetime.utcnow()
        }
        result = purchase_items_collection.insert_one(item)
        item['_id'] = str(result.inserted_id)
        return jsonify({'message': 'Item added successfully', 'item': item}), 201
    except Exception as e:
        return jsonify({'error': f"Failed to add item: {str(e)}"}), 500

@app.route('/api/purchase_items/<id>', methods=['PUT'])
def update_purchase_item(id):
    try:
        data = request.json
        if not data:
            return jsonify({'error': 'No input data provided'}), 400
        update_fields = {}
        if 'name' in data:
            update_fields['name'] = data['name']
        if 'mainUnit' in data:
            update_fields['mainUnit'] = data['mainUnit']
        if 'subUnit' in data:
            update_fields['subUnit'] = data['subUnit']
        if 'conversionFactor' in data:
            update_fields['conversionFactor'] = float(data['conversionFactor'])
        if not update_fields:
            return jsonify({'error': 'No fields to update'}), 400
        result = purchase_items_collection.update_one({'_id': ObjectId(id)}, {'$set': update_fields})
        if result.modified_count == 0:
            return jsonify({'error': 'Item not found or no changes made'}), 404
        return jsonify({'message': 'Item updated successfully'}), 200
    except ValueError:
        return jsonify({'error': 'Invalid input data'}), 400
    except Exception as e:
        return jsonify({'error': f"Failed to update item: {str(e)}"}), 500

@app.route('/api/purchase_items/<id>', methods=['DELETE'])
def delete_purchase_item(id):
    try:
        result = purchase_items_collection.delete_one({'_id': ObjectId(id)})
        if result.deleted_count == 0:
            return jsonify({'error': 'Item not found'}), 404
        return jsonify({'message': 'Item deleted successfully'}), 200
    except ValueError:
        return jsonify({'error': 'Invalid item ID'}), 400
    except Exception as e:
        return jsonify({'error': f"Failed to delete item: {str(e)}"}), 500

# --- Suppliers Routes ---
@app.route('/api/suppliers', methods=['GET'])
def get_suppliers():
    try:
        suppliers = list(suppliers_collection.find())
        return jsonify([document_to_dict(supplier) for supplier in suppliers]), 200
    except Exception as e:
        return jsonify({'error': f"Failed to fetch suppliers: {str(e)}"}), 500

@app.route('/api/suppliers', methods=['POST'])
def add_supplier():
    try:
        data = request.json
        required_fields = ['name', 'shopName', 'address', 'phone', 'email']
        if not all(key in data for key in required_fields):
            return jsonify({'error': 'Missing required fields'}), 400
        if not all(data[key] for key in required_fields):
            return jsonify({'error': 'Invalid input data'}), 400
        
        supplier = {
            'name': data['name'],
            'shopName': data['shopName'],
            'address': data['address'],
            'phone': data['phone'],
            'email': data['email'],
            'created_at': datetime.utcnow()
        }
        result = suppliers_collection.insert_one(supplier)
        supplier['_id'] = str(result.inserted_id)
        return jsonify({'message': 'Supplier added successfully', 'supplier': supplier}), 201
    except Exception as e:
        return jsonify({'error': f"Failed to add supplier: {str(e)}"}), 500

@app.route('/api/suppliers/<id>', methods=['PUT'])
def update_supplier(id):
    try:
        data = request.json
        if not data:
            return jsonify({'error': 'No input data provided'}), 400
        update_fields = {}
        if 'name' in data:
            update_fields['name'] = data['name']
        if 'shopName' in data:
            update_fields['shopName'] = data['shopName']
        if 'address' in data:
            update_fields['address'] = data['address']
        if 'phone' in data:
            update_fields['phone'] = data['phone']
        if 'email' in data:
            update_fields['email'] = data['email']
        if not update_fields:
            return jsonify({'error': 'No fields to update'}), 400
        result = suppliers_collection.update_one({'_id': ObjectId(id)}, {'$set': update_fields})
        if result.modified_count == 0:
            return jsonify({'error': 'Supplier not found or no changes made'}), 404
        return jsonify({'message': 'Supplier updated successfully'}), 200
    except ValueError:
        return jsonify({'error': 'Invalid input data'}), 400
    except Exception as e:
        return jsonify({'error': f"Failed to update supplier: {str(e)}"}), 500

@app.route('/api/suppliers/<id>', methods=['DELETE'])
def delete_supplier(id):
    try:
        result = suppliers_collection.delete_one({'_id': ObjectId(id)})
        if result.deleted_count == 0:
            return jsonify({'error': 'Supplier not found'}), 404
        return jsonify({'message': 'Supplier deleted successfully'}), 200
    except ValueError:
        return jsonify({'error': 'Invalid supplier ID'}), 400
    except Exception as e:
        return jsonify({'error': f"Failed to delete supplier: {str(e)}"}), 500


# --- Purchase Orders Routes ---
@app.route('/api/purchase_orders', methods=['GET'])
def get_purchase_orders():
    try:
        orders = list(purchase_orders_collection.find())
        return jsonify([document_to_dict(order) for order in orders]), 200
    except Exception as e:
        return jsonify({'error': f"Failed to fetch purchase orders: {str(e)}"}), 500

@app.route('/api/purchase_orders', methods=['POST'])
def add_purchase_order():
    try:
        data = request.get_json()
        required_fields = ['supplierId', 'date', 'items']
        if not all(key in data for key in required_fields):
            return jsonify({'error': 'Missing required fields'}), 400
        if not data['supplierId'] or not data['date'] or not isinstance(data['items'], list) or not data['items']:
            return jsonify({'error': 'Invalid input data'}), 400
        
        supplier = suppliers_collection.find_one({'_id': ObjectId(data['supplierId'])})
        if not supplier:
            return jsonify({'error': 'Supplier not found'}), 404
        
        for item in data['items']:
            if not all(key in item for key in ['itemId', 'quantity', 'unit']):
                return jsonify({'error': 'Invalid item data'}), 400
            if not item['itemId'] or float(item['quantity']) <= 0 or item['unit'] not in ['main', 'sub']:
                return jsonify({'error': 'Invalid item quantity, ID, or unit'}), 400
            if not purchase_items_collection.find_one({'_id': ObjectId(item['itemId'])}):
                return jsonify({'error': f"Item {item['itemId']} not found"}), 404

        # Generate unique ID using first 8 characters of a UUID
        order_id = f"PO-{str(uuid.uuid4())[:8].upper()}"
        order = {
            'id': order_id,
            'supplierId': data['supplierId'],
            'supplierName': supplier['name'],
            'supplierEmail': supplier['email'],
            'date': datetime.strptime(data['date'], '%Y-%m-%d'),
            'items': data['items'],
            'status': 'Pending',
            'created_at': datetime.utcnow()
        }
        result = purchase_orders_collection.insert_one(order)
        order['_id'] = str(result.inserted_id)
        return jsonify({'message': 'Purchase Order created successfully', 'order': document_to_dict(order)}), 201
    except ValueError as e:
        return jsonify({'error': f"Invalid data format: {str(e)}"}), 400
    except Exception as e:
        return jsonify({'error': f"Failed to create purchase order: {str(e)}"}), 500

@app.route('/api/purchase_orders/<id>', methods=['DELETE'])
def delete_purchase_order(id):
    try:
        result = purchase_orders_collection.delete_one({'id': id})
        if result.deleted_count == 0:
            return jsonify({'error': 'Purchase Order not found'}), 404
        return jsonify({'message': 'Purchase Order deleted successfully'}), 200
    except Exception as e:
        return jsonify({'error': f"Failed to delete purchase order: {str(e)}"}), 500

# --- Purchase Receipts Routes ---
@app.route('/api/purchase_receipts', methods=['GET'])
def get_purchase_receipts():
    try:
        receipts = list(purchase_receipts_collection.find())
        return jsonify([document_to_dict(receipt) for receipt in receipts]), 200
    except Exception as e:
        return jsonify({'error': f"Failed to fetch purchase receipts: {str(e)}"}), 500

@app.route('/api/purchase_receipts', methods=['POST'])
def add_purchase_receipt():
    try:
        data = request.get_json()
        required_fields = ['poId', 'date', 'items']
        if not all(key in data for key in required_fields):
            return jsonify({'error': 'Missing required fields'}), 400
        if not data['poId'] or not data['date'] or not isinstance(data['items'], list) or not data['items']:
            return jsonify({'error': 'Invalid input data'}), 400
        
        po = purchase_orders_collection.find_one({'id': data['poId']})
        if not po:
            return jsonify({'error': 'Purchase Order not found'}), 404
        
        for item in data['items']:
            if not all(key in item for key in ['itemId', 'quantity', 'unit', 'status']):
                return jsonify({'error': 'Invalid item data'}), 400
            if not item['itemId'] or float(item['quantity']) <= 0 or item['status'] not in ['Accepted', 'Rejected'] or item['unit'] not in ['main', 'sub']:
                return jsonify({'error': 'Invalid item quantity, ID, unit, or status'}), 400
            if not purchase_items_collection.find_one({'_id': ObjectId(item['itemId'])}):
                return jsonify({'error': f"Item {item['itemId']} not found"}), 404

        receipt_id = f"PR-{str(uuid.uuid4())[:8].upper()}"
        receipt = {
            'id': receipt_id,
            'poId': data['poId'],
            'date': datetime.strptime(data['date'], '%Y-%m-%d'),
            'items': data['items'],
            'created_at': datetime.utcnow()
        }
        result = purchase_receipts_collection.insert_one(receipt)
        receipt['_id'] = str(result.inserted_id)
        return jsonify({'message': 'Purchase Receipt created successfully', 'receipt': document_to_dict(receipt)}), 201
    except ValueError as e:
        return jsonify({'error': f"Invalid data format: {str(e)}"}), 400
    except Exception as e:
        return jsonify({'error': f"Failed to create purchase receipt: {str(e)}"}), 500

@app.route('/api/purchase_receipts/<id>', methods=['DELETE'])
def delete_purchase_receipt(id):
    try:
        result = purchase_receipts_collection.delete_one({'id': id})
        if result.deleted_count == 0:
            return jsonify({'error': 'Purchase Receipt not found'}), 404
        return jsonify({'message': 'Purchase Receipt deleted successfully'}), 200
    except Exception as e:
        return jsonify({'error': f"Failed to delete purchase receipt: {str(e)}"}), 500

# --- Purchase Invoices Routes ---
@app.route('/api/purchase_invoices', methods=['GET'])
def get_purchase_invoices():
    try:
        invoices = list(purchase_invoices_collection.find())
        return jsonify([document_to_dict(invoice) for invoice in invoices]), 200
    except Exception as e:
        return jsonify({'error': f"Failed to fetch purchase invoices: {str(e)}"}), 500

@app.route('/api/purchase_invoices', methods=['POST'])
def add_purchase_invoice():
    try:
        data = request.get_json()
        required_fields = ['poId', 'date', 'supplier', 'items']
        if not all(key in data for key in required_fields):
            return jsonify({'error': 'Missing required fields'}), 400
        if not data['poId'] or not data['date'] or not data['supplier'] or not isinstance(data['items'], list) or not data['items']:
            return jsonify({'error': 'Invalid input data'}), 400
        
        po = purchase_orders_collection.find_one({'id': data['poId']})
        if not po:
            return jsonify({'error': 'Purchase Order not found'}), 404
        
        if data.get('prId'):
            pr = purchase_receipts_collection.find_one({'id': data['prId']})
            if not pr:
                return jsonify({'error': 'Purchase Receipt not found'}), 404
        
        for item in data['items']:
            if not all(key in item for key in ['itemId', 'quantity', 'unit', 'rate', 'tax']):
                return jsonify({'error': 'Invalid item data'}), 400
            if not item['itemId'] or float(item['quantity']) <= 0 or float(item['rate']) <= 0 or float(item['tax']) < 0 or item['unit'] not in ['main', 'sub']:
                return jsonify({'error': 'Invalid item quantity, rate, tax, or unit'}), 400
            if not purchase_items_collection.find_one({'_id': ObjectId(item['itemId'])}):
                return jsonify({'error': f"Item {item['itemId']} not found"}), 404

        invoice_id = f"PI-{str(uuid.uuid4())[:8].upper()}"
        invoice = {
            'id': invoice_id,
            'poId': data['poId'],
            'prId': data.get('prId', ''),
            'date': datetime.strptime(data['date'], '%Y-%m-%d'),
            'supplier': data['supplier'],
            'items': data['items'],
            'created_at': datetime.utcnow()
        }
        result = purchase_invoices_collection.insert_one(invoice)
        invoice['_id'] = str(result.inserted_id)
        return jsonify({'message': 'Purchase Invoice created successfully', 'invoice': document_to_dict(invoice)}), 201
    except ValueError as e:
        return jsonify({'error': f"Invalid data format: {str(e)}"}), 400
    except Exception as e:
        return jsonify({'error': f"Failed to create purchase invoice: {str(e)}"}), 500

@app.route('/api/purchase_invoices/<id>', methods=['DELETE'])
def delete_purchase_invoice(id):
    try:
        result = purchase_invoices_collection.delete_one({'id': id})
        if result.deleted_count == 0:
            return jsonify({'error': 'Purchase Invoice not found'}), 404
        return jsonify({'message': 'Purchase Invoice deleted successfully'}), 200
    except Exception as e:
        return jsonify({'error': f"Failed to delete purchase invoice: {str(e)}"}), 500




# Serve React react
@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def serve_frontend(path):
    """Serve the React frontend from the dist folder."""
    if path != "" and os.path.exists(os.path.join(STATIC_DIR, path)):
        return send_from_directory(STATIC_DIR, path)
    if os.path.exists(os.path.join(STATIC_DIR, 'index.html')):
        return send_from_directory(STATIC_DIR, 'index.html')
    logger.warning(f"Frontend file not found: {path}")
    return jsonify({"error": "Frontend not found"}), 404

# Start the server
def start_scheduler():
    pass  # Add your scheduler logic here if required

if __name__ == '__main__':
    start_scheduler()  # Start the backup scheduler
    # Use Waitress for production, Flask's built-in server for development
    if getattr(sys, 'frozen', False):
        logger.info("Running as frozen executable, using Waitress")
        waitress.serve(app, host='0.0.0.0', port=5000, threads=1, _quiet=True)
    else:
        logger.info("Running in development mode, using Flask")
        app.run(host='0.0.0.0', port=5000, debug=True)